from flask import Blueprint, render_template, url_for, request, session, flash, redirect, make_response, jsonify, current_app
from pymongo import ASCENDING
from datetime import datetime, timedelta
from utils import get_mongo_client, get_db_and_fs, send_async_email
import pytz
from bson.objectid import ObjectId
from io import BytesIO
from openpyxl import Workbook, load_workbook

logs = Blueprint('logs_route', __name__)

utc = pytz.UTC

def rename_fourth_field(doc):
    keys = list(doc.keys())
    if len(keys) >= 4:
        # Fourth field's key (position 3 in zero-indexed list)
        fourth_key = keys[3]
        doc['details'] = doc.pop(fourth_key)
    if 'timestamp' in doc and isinstance(doc['timestamp'], datetime):
        doc['timestamp'] = doc['timestamp'].strftime('%Y-%m-%d %H:%M')
    return doc

def convert_to_eat(timestamp):
    # Parse the timestamp (assuming it's in ISO 8601 format)
    utc_dt = datetime.fromisoformat(timestamp)
    # Define the UTC and EAT timezones
    utc = pytz.utc
    eat = pytz.timezone('Africa/Nairobi')
    # Localize the datetime to UTC
    utc_dt = utc.localize(utc_dt)
    # Convert the datetime to EAT
    eat_dt = utc_dt.astimezone(eat)
    # Return the formatted datetime string
    return eat_dt.strftime('%Y-%m-%d %H:%M')

##AUDIT LOGS
@logs.route('/view-audit-logs')
def view_audit_logs():
    db, fs = get_db_and_fs()
    username = session.get('login_username')
    if username is None:
        flash('Login first', 'error')
        return redirect('/')
    else:
        company = db.registered_managers.find_one({'username': username})
        if 'dp' in company:
            dp_str = company['dp']
        else:
            dp_str = None
        usernames = db.registered_managers.find({'company_name': company['company_name']}, {'username': 1})
        renamed_logs = []
        is_manager = db.managers.find_one({'manager_email': company['email'], 'name':company['company_name']})
        if is_manager:
            for user in usernames:
                audit_logs = db.audit_logs.find({'user': user['username']})
                for log in audit_logs:
                    renamed_log = rename_fourth_field(log)
                    timestamp = log.get('timestamp')
                    log['timestamp'] = convert_to_eat(timestamp)
                    renamed_logs.append(renamed_log)
            sorted_logs = sorted(renamed_logs, key=lambda x: x["timestamp"], reverse=True)
            logs_first_40 = sorted_logs[:40]
        else:
            renamed_logs = []
        return render_template('audit logs.html', audit_logs=logs_first_40, dp=dp_str)

# Function to rename the fourth field to 'details'
def format_time(doc):
    if 'timestamp' in doc and isinstance(doc['timestamp'], datetime):
        doc['timestamp'] = doc['timestamp'].strftime('%Y-%m-%d %H:%M')
    return doc

##LOGIN HISTORY
@logs.route('/view-login-history')
def view_login_history():
    db, fs = get_db_and_fs()
    username = session.get('login_username')
    if username is None:
        flash('Login first', 'error')
        return redirect('/')
    else:
        company = db.registered_managers.find_one({'username': username})
        if 'dp' in company:
            dp_str = company['dp']
        else:
            dp_str = None
        usernames = db.registered_managers.find({'company_name': company['company_name']}, {'username': 1})
        logindata = []
        is_manager = db.managers.find_one({'manager_email': company['email'], 'name':company['company_name']})
        if is_manager:
            for user in usernames:
                login_info = db.logged_in_data.find({'username': user['username']})
                for login in login_info:
                    formated_time = format_time(login)
                    timestamp = login.get('timestamp')
                    login['timestamp'] = convert_to_eat(timestamp)
                    logindata.append(formated_time)
            sorted_logins = sorted(logindata, key=lambda x: x["timestamp"], reverse=True)
            logindata_first_40 = sorted_logins[:40]
        else:
            logindata = []
        return render_template('login history.html', logindata=logindata_first_40, dp=dp_str)

###DOANLOAD AUDIT DATA   
@logs.route('/download-audit-logs', methods=["POST"])
def download_audit_logs():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/')
    
    startdate_on_str = request.form.get("startdate")
    enddate_on_str = request.form.get("enddate")
    startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
    enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')
    
    company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'address': 0, 'password': 0, 'auth': 0, 'dark_mode': 0})
    if not company:
        flash('Company not found', 'error')
        return redirect('/')
    
    usernames = db.registered_managers.find({'company_name': company['company_name']}, {'username': 1})
    renamed_logs = []

    for user in usernames:
        audit_logs = list(db.audit_logs.find({
            'user': user['username'],
            'timestamp': {'$gte': startdate, '$lte': enddate}
        }))
        if audit_logs:
            for log in audit_logs:
                renamed_log = rename_fourth_field(log)
                timestamp = log.get('timestamp')
                renamed_log['timestamp'] = convert_to_eat(timestamp)

                # Convert non-serializable types to strings
                for key, value in log.items():
                    if isinstance(value, ObjectId):
                        log[key] = str(value)
                    elif isinstance(value, (bytes, bytearray)):
                        log[key] = value.decode('utf-8')

                renamed_logs.append(renamed_log)
    
    # Sort logs by timestamp
    sorted_logs = sorted(renamed_logs, key=lambda x: x["timestamp"], reverse=True)

    # Create an Excel workbook
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    # Define headers based on sorted_logs keys
    headers = sorted_logs[0].keys() if sorted_logs else []
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    for r_idx, log in enumerate(sorted_logs, start=2):
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=log.get(header, ''))

    # Save the workbook to a BytesIO buffer
    output.seek(0)
    wb.save(output)
    wb.close()

    # Set the buffer position to the beginning
    output.seek(0)
    excel_data = output.read()

    # Create the response with the Excel file
    response = make_response(excel_data)
    response.headers['Content-Disposition'] = f"attachment; filename={company['company_name']}_audit_logs_{startdate_on_str}_{enddate_on_str}.xlsx"
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response

###DOANLOAD LOGIN DATA   
@logs.route('/download-login-data', methods=["POST"])
def download_login_data():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/')
    
    startdate_on_str = request.form.get("startdate")
    enddate_on_str = request.form.get("enddate")
    startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
    enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')
    company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'address': 0, 'password': 0, 'auth': 0, 'dark_mode': 0})
    usernames = db.registered_managers.find({'company_name': company['company_name']}, {'username': 1})
    logindata = []
    for user in usernames:
        login_info = list(db.logged_in_data.find({'username': user['username'], 'timestamp': {'$gte': startdate, '$lte': enddate}}))
        if login_info:
            for login in login_info:
                formated_time = format_time(login)
                timestamp = login.get('timestamp')
                login['timestamp'] = convert_to_eat(timestamp)

                # Convert non-serializable types to strings
                for key, value in login.items():
                    if isinstance(value, ObjectId):
                        login[key] = str(value)
                    elif isinstance(value, (bytes, bytearray)):
                        login[key] = value.decode('utf-8')
                        
                logindata.append(formated_time)
    
    # Sort logins
    sorted_logins = sorted(logindata, key=lambda x: x["timestamp"], reverse=True)

    # Create an Excel workbook
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Login Data"
    
    # Define headers based on sorted_logs keys
    headers = sorted_logins[0].keys() if sorted_logins else []
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data rows
    for r_idx, log in enumerate(sorted_logins, start=2):
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=log.get(header, ''))
    
    # Save the workbook to a BytesIO buffer
    output.seek(0)
    wb.save(output)
    wb.close()

    # Set the buffer position to the beginning
    output.seek(0)
    excel_data = output.read()

    # Create the response
    response = make_response(excel_data)
    response.headers['Content-Disposition'] = f"attachment; filename={company['company_name']}_login_data_{startdate_on_str}_{enddate_on_str}.xlsx"
    response.headers['Content-Type'] = 'application/zip'

    return response