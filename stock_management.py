from flask import Blueprint, render_template, url_for, request, session, flash, redirect, make_response, jsonify, send_from_directory, current_app
from pymongo import ASCENDING
from datetime import datetime, timedelta
import bcrypt
from utils import get_mongo_client, get_db_and_fs, send_async_email
from flask_mail import Message
import threading
from docx import Document
import calendar
import pytz
import pandas as pd 
from io import BytesIO
import json
from bson.objectid import ObjectId
import cv2
import numpy as np
import io
import base64
import random
import os
from werkzeug.utils import secure_filename
from gridfs import GridFS
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import ZipFile
import tempfile
import string
import qrcode
import time
from docx2pdf import convert
import PyPDF2
import gc
from collections import defaultdict
import barcode
from PIL import Image, ImageDraw, ImageFont
from barcode.writer import ImageWriter

stockManagement = Blueprint('stockManagement_route', __name__)

def generate_random_product_id(length=8):
    characters = string.ascii_letters + string.digits  # a-z, A-Z, 0-9
    return ''.join(random.choice(characters) for _ in range(length))

@stockManagement.route('/add new stock page')
def add_new_stock_page():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            if session.get('add_stock') == "no":
                flash("You do not have rights to add stock","error")
                return redirect('/stock-overview')
            else:
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                return render_template('add new stock.html', dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/update existing stock')
def update_existing_stock():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            if session.get('update_stock') == "no":
                flash("You do not have rights to update stock","error")
                return redirect('/stock-overview')
            else:
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if not company:
                    flash('Company not found', 'error')
                    return redirect('/manager login page')
                
                dp_str = company.get('dp')
                items_to_update = []
                available_items = db.inventories.find({'company_name': company['company_name']})
                for item in available_items:
                    item_details = {
                        'itemName': item['itemName'],
                        'available_quantity': item.get('available_quantity', ''),
                        'unitOfMeasurement': item.get('unitOfMeasurement', ''),
                        'unitPrice': item.get('unitPrice','')
                    }
                    items_to_update.append(item_details)
                
                items_to_update = sorted(items_to_update, key=lambda x: x['itemName'])

                return render_template('update existing stock.html', dp=dp_str, items_to_update=items_to_update)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/update sales page')
def update_sales_page():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            if session.get('update_sales') == "no":
                flash("You do not have rights to update sales","error")
                return redirect('/stock-overview')
            else:
                company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'address': 0, 'password': 0, 'auth': 0, 'dark_mode': 0})
                if not company:
                    flash('Company not found', 'error')
                    return redirect('/manager login page')
                
                dp_str = company.get('dp')
                available_itemNames = []
                available_items = db.inventories.find({'company_name': company['company_name']})
                for item in available_items:
                    if item.get('available_quantity', 0) > 0:
                        item_dict = {
                            'itemName': item.get('itemName', ''),
                            'available_quantity': item.get('available_quantity', ''),
                            'unitOfMeasurement': item.get('unitOfMeasurement', '')
                        }
                        
                        # Add the 'sellingPrice' field if it exists in the item
                        if 'selling_price' in item:
                            item_dict['selling_price'] = item['selling_price']

                        # Append the item_dict to the available_itemNames list
                        available_itemNames.append(item_dict)

                # Sort the available_itemNames list in alphabetical order by 'itemName'
                available_itemNames = sorted(available_itemNames, key=lambda x: x['itemName'])

                return render_template('update sales page.html', dp=dp_str, available_itemNames=available_itemNames)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/scan bar code page')
def scan_bar_code_page():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            if session.get('update_sales') == "no":
                flash("You do not have rights to update sales","error")
                return redirect('/stock-overview')
            else:
                company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'address': 0, 'password': 0, 'auth': 0, 'dark_mode': 0})
                if not company:
                    flash('Company not found', 'error')
                    return redirect('/manager login page')
                
                dp_str = company.get('dp')
                available_itemNames = []
                available_items = db.inventories.find({'company_name': company['company_name']})
                for item in available_items:
                    if item.get('available_quantity', 0) > 0:
                        item_dict = {
                            'product_id': item.get('product_id', ''),
                            'itemName': item.get('itemName', ''),
                            'available_quantity': item.get('available_quantity', ''),
                        }
                        
                        # Add the 'sellingPrice' field if it exists in the item
                        if 'selling_price' in item:
                            item_dict['selling_price'] = item['selling_price']

                        # Append the item_dict to the available_itemNames list
                        available_itemNames.append(item_dict)

                # Sort the available_itemNames list in alphabetical order by 'itemName'
                available_itemNames = sorted(available_itemNames, key=lambda x: x['itemName'])
                return render_template('scan bar code.html', dp=dp_str,available_itemNames=available_itemNames)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/generate product bar codes page')
def generate_bar_codes():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            if session.get('update_stock') == "no":
                flash("You do not have rights to update stock","error")
                return redirect('/stock-overview')
            else:
                company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'address': 0, 'password': 0, 'auth': 0, 'dark_mode': 0})
                if not company:
                    flash('Company not found', 'error')
                    return redirect('/manager login page')
                
                dp_str = company.get('dp')
                available_itemNames = []
                available_items = db.inventories.find({'company_name': company['company_name']})
                for item in available_items:
                    available_itemNames.append({
                        'itemName': item.get('itemName', '')
                    })

                # Sort the available_itemNames list in alphabetical order by 'itemName'
                available_itemNames = sorted(available_itemNames, key=lambda x: x['itemName'])

                return render_template('generate barcodes.html', dp=dp_str, available_itemNames=available_itemNames)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/update production activity')
def update_production_activity():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            if session.get('inhouse') == "no":
                flash("You do not have rights to update inhouse information","error")
                return redirect('/stock-overview')
            else:
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                
                available_itemNames = []
                available_items = db.inventories.find({'company_name': company['company_name']})
                for item in available_items:
                    if item.get('available_quantity', 0) > 0:
                        available_itemNames.append({
                            'itemName': item.get('itemName', ''),  # Provide a default value
                            'available_quantity': item.get('available_quantity', ''),
                            'unitOfMeasurement': item.get('unitOfMeasurement', '')  # Provide a default value
                        })
                    
                available_itemNames = sorted(available_itemNames, key=lambda x: x['itemName'])
                return render_template('update production.html', dp=dp_str, available_itemNames=available_itemNames)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/update inhouse use page')
def update_inhouse_use_page():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            if session.get('inhouse') == "no":
                flash("You do not have rights to update inhouse information","error")
                return redirect('/stock-overview')
            else:
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None

                available_itemNames = []
                available_items = db.inventories.find({'company_name': company['company_name']})
                for item in available_items:
                    if item.get('available_quantity', 0) > 0:
                        available_itemNames.append({
                            'itemName': item.get('itemName', ''),  # Provide a default value
                            'available_quantity': item.get('available_quantity', ''),
                            'unitOfMeasurement': item.get('unitOfMeasurement', '')  # Provide a default value
                        })

                available_itemNames = sorted(available_itemNames, key=lambda x: x['itemName'])
                return render_template('update inhouse use.html', dp=dp_str, available_itemNames=available_itemNames)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/add-new-stock', methods=['POST'])
def add_new_stock():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
                
            all_items = request.json.get('items', [])  # Access the JSON data sent from the client
            skipped_items = []  # List to hold names of items that were not added
            added_items = []  # List to hold names of items that were successfully added
            timestamp = datetime.now()
            generated_ids = set()

            for item in all_items:
                item['itemName'] = item.get('itemName', '').strip()
                
                try:
                    # Convert 'quantity' and 'unitPrice' to floats
                    item['quantity'] = float(item.get('quantity', 0))
                    item['available_quantity'] = item['quantity']
                    item['unitPrice'] = float(item.get('unitPrice', 0))
                    item['stockDate'] = datetime.strptime(item.get('stockDate', ''), '%Y-%m-%d')

                    selling_price = item.get('selling_price')
                    if selling_price:
                        item['selling_price'] = float(selling_price)
                    else:
                        item.pop('selling_price', None)

                    # Add 'totalPrice' field which is 'unitPrice' * 'quantity'
                    item['totalPrice'] = item['unitPrice'] * item['quantity']
                    item['company_name'] = company.get('company_name', '')
                    item['timestamp'] = timestamp
                    item['oldTotalPrice'] = item['totalPrice']
                    item['cumulativeOldPrices'] = item['totalPrice']
                    
                    new_product_id = generate_random_product_id()
                    while new_product_id in generated_ids or db.inventories.find_one({'company_name': item['company_name'],'product_id': new_product_id}):
                        new_product_id = generate_random_product_id()
                    # Add the new unique product_id to the set
                    generated_ids.add(new_product_id)
                    item['product_id'] = new_product_id

                    # Check if the item already exists in the database
                    existing_item = db.inventories.find_one({
                        'itemName': item['itemName'],
                        'company_name': item['company_name']
                    })

                    if existing_item:
                        skipped_items.append(item['itemName'])  # Add the name of the skipped item
                        continue  # Skip this iteration and don't add the existing item

                    # Insert the new stock entry into MongoDB
                    db.inventories.insert_one(item)
                    db.audit_logs.insert_one({
                        'user': login_data,
                        'Activity': 'Added new item to stock',
                        'Item': item['itemName'],
                        'timestamp': datetime.now()
                    })
                    added_items.append(item['itemName'])
                except (ValueError, TypeError) as e:
                    # Log or handle the exception as needed
                    flash(f"Error processing item {item.get('itemName', 'unknown')}: {e}", 'error')
                    skipped_items.append(item.get('itemName', 'unknown'))

            message = ""
            if added_items:
                message += '. The following items were added: ' + ', '.join(added_items)
                flash(message, 'success')
            if skipped_items:
                message_skipped = 'The following items were not added because they already exist: ' + ', '.join(skipped_items)
                flash(message_skipped, 'error')

            return jsonify({'redirect': url_for('stockManagement_route.add_new_stock_page')})
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@stockManagement.route('/update-new-stock', methods=['POST'])
def update_new_stock():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
                
            all_items = request.json.get('items', [])  # Access the JSON data sent from the client
            timestamp = datetime.now()
            generated_ids = set()

            for item in all_items:
                item['itemName'] = item.get('itemName', '').strip()

                try:
                    # Convert 'quantity' and 'unitPrice' to floats
                    item['quantity'] = float(item.get('quantity', 0))
                    item['unitPrice'] = float(item.get('unitPrice', 0))
                    item['stockDate'] = datetime.strptime(item.get('stockDate', ''), '%Y-%m-%d')
                    item['company_name'] = company.get('company_name', '')
                    item['status'] = "updated stock"
                    item['timestamp'] = timestamp

                    selling_price = item.get('selling_price')
                    if selling_price:
                        item['selling_price'] = float(selling_price)
                    else:
                        item.pop('selling_price', None)

                    # Check if the item already exists in the database
                    existing_item = db.inventories.find_one({
                        'itemName': item['itemName'],
                        'company_name': item['company_name']
                    })

                    if existing_item:
                        if 'product_id' in existing_item:
                            item['product_id'] = existing_item['product_id']
                        else:
                            new_product_id = generate_random_product_id()
                            while new_product_id in generated_ids or db.inventories.find_one({'company_name': existing_item['company_name'],'product_id': new_product_id}):
                                new_product_id = generate_random_product_id()
                            # Add the new unique product_id to the set
                            generated_ids.add(new_product_id)
                            item['product_id'] = new_product_id

                        selling_price = item.get('selling_price')

                        if selling_price:
                            item['selling_price'] = float(item['selling_price'])
                        else:
                            if 'selling_price' in existing_item:
                                item['selling_price'] = existing_item['selling_price']
                            else:
                                item['selling_price'] = 0
                        if 'available_quantity' in existing_item:
                            if existing_item['available_quantity'] > 0:
                                # Add 'totalPrice' field which is 'unitPrice' * 'quantity'
                                item['totalPrice'] = item['quantity'] * item['unitPrice']
                                item['unitOfMeasurement'] = existing_item.get('unitOfMeasurement', '')
                                item['oldTotalPrice'] = existing_item.get('oldTotalPrice', existing_item.get('totalPrice'))
                                item['cumulativeOldPrices'] = existing_item.get('cumulativeOldPrices', existing_item.get('totalPrice')) + item['totalPrice']
                                item['oldUnitPrice'] = existing_item.get('unitPrice', 0)
                                new_available_quantity = existing_item['available_quantity'] + item['quantity']
                                item['available_quantity'] = new_available_quantity
                            else:
                                new_available_quantity = existing_item['available_quantity'] + item['quantity']
                                item['available_quantity'] = new_available_quantity
                                item['totalPrice'] = item['quantity'] * item['unitPrice']
                                item['unitOfMeasurement'] = existing_item.get('unitOfMeasurement', '')
                                item['oldTotalPrice'] = existing_item.get('oldTotalPrice', existing_item.get('totalPrice'))
                                item['cumulativeOldPrices'] = existing_item.get('cumulativeOldPrices', existing_item.get('totalPrice')) + item['totalPrice']
                        else:
                            new_available_quantity = item['quantity']
                            item['available_quantity'] = new_available_quantity
                            item['totalPrice'] = item['quantity'] * item['unitPrice']
                            item['unitOfMeasurement'] = existing_item.get('unitOfMeasurement', '')
                            item['oldTotalPrice'] = existing_item.get('oldTotalPrice', existing_item.get('totalPrice'))
                            item['cumulativeOldPrices'] = existing_item.get('cumulativeOldPrices', existing_item.get('totalPrice')) + item['totalPrice']

                        # Insert the updated stock entry into MongoDB
                        db.inventories.insert_one(item)
                        db.audit_logs.insert_one({
                            'user': login_data,
                            'Activity': 'Updated item in stock',
                            'Item': item['itemName'],
                            'timestamp': datetime.now()
                        })
                        db.inventories.delete_one({'_id': existing_item['_id']})
                        existing_item.pop('_id', None)
                        db.old_inventories.insert_one(existing_item)
                    else:
                        # Handle case where item does not exist if necessary
                        flash(f"Item {item['itemName']} does not exist.", 'error')
                except (ValueError, TypeError) as e:
                    flash(f"Error processing item {item.get('itemName', 'unknown')}: {e}", 'error')

            flash('Stock updated successfully', 'success')
            return jsonify({'redirect': url_for('stockManagement_route.update_existing_stock')})
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@stockManagement.route('/update-sale', methods=['POST'])
def update_sale():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
                
            all_items = request.json.get('items', [])  # Access the JSON data sent from the client
            out_of_stock_items = []
            over_quantified = []
            timestamp = datetime.now()
            updates = 0
            for item in all_items:
                try:
                    item['quantity'] = float(item.get('quantity', 0))
                    item['unitPrice'] = float(item.get('unitPrice', 0))
                    if item['saleDate'] == "null":
                        item['saleDate'] = timestamp
                    else:
                        item['saleDate'] = datetime.strptime(item.get('saleDate', ''), '%Y-%m-%d')
                    item['company_name'] = company.get('company_name', '')
                    item['timestamp'] = timestamp

                    existing_item = db.inventories.find_one({
                        'itemName': item['itemName'],
                        'company_name': company['company_name']
                    })

                    if existing_item:
                        updates = 1
                        if 'available_quantity' in existing_item:
                            if existing_item['available_quantity'] <= 0:
                                out_of_stock_items.append(item['itemName'])
                                continue
                            if item['quantity'] > existing_item['available_quantity']:
                                over_quantified.append(item['itemName'])
                                continue
                            revenue = item['quantity'] * item['unitPrice']
                            available_quantity = existing_item['available_quantity'] - item['quantity']
                            item['revenue'] = revenue
                            item['stockDate'] = existing_item['stockDate']
                        else:
                            if item['quantity'] > existing_item['quantity']:
                                over_quantified.append(item['itemName'])
                                continue
                            revenue = item['quantity'] * item['unitPrice']
                            available_quantity = existing_item['quantity'] - item['quantity']
                            item['revenue'] = revenue
                            item['stockDate'] = existing_item['stockDate']
                        
                        item['stock_id'] = existing_item['_id']

                        db.stock_sales.insert_one(item)
                        db.audit_logs.insert_one({
                            'user': login_data,
                            'Activity': 'Added a new sale',
                            'Item': item['itemName'],
                            'timestamp': datetime.now()
                        })
                        db.inventories.update_one({'_id': existing_item['_id']}, {'$set': {'available_quantity': available_quantity}})
                    else:
                        flash(f"Item {item['itemName']} does not exist.", 'error')
                except (ValueError, TypeError) as e:
                    flash(f"Error processing item {item.get('itemName', 'unknown')}: {e}", 'error')
            
            if updates ==1:
                message = 'Sales updated successfully'
                flash(message, 'success')

            if out_of_stock_items:
                flash(f'The following items are out of stock: {", ".join(out_of_stock_items)}', 'error')
            if over_quantified:
                flash(f'Enter smaller quantities for the following items: {", ".join(over_quantified)}', 'error')

            return jsonify({'redirect': url_for('stockManagement_route.update_sales_page')})
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/store_scanned_sale', methods=['GET', 'POST'])
def store_scanned_sale():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    
    company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                        'password': 0, 'auth': 0, 'dark_mode': 0})

    scanned_items_json = request.form.get('scanned_items')
    if scanned_items_json:
        scanned_items = json.loads(scanned_items_json)
        success_messages = []
        error_messages = []
        for item in scanned_items:
            product_id = item['product_id']
            sold_quantity = item['sold_quantity']
            sold_quantity = float(sold_quantity)
        
            existing_item = db.inventories.find_one({'company_name': company['company_name'], 'product_id': product_id})

            if existing_item:
                timestamp = datetime.now()
                if 'available_quantity' in existing_item:
                    if existing_item['available_quantity'] > 0:
                        if existing_item['available_quantity'] >= sold_quantity:
                            available_quantity = existing_item['available_quantity'] - sold_quantity
                            stockDate = existing_item['stockDate']
                        else:
                            error_messages.append(f'Item {existing_item["itemName"]} has insufficient stock.')
                            continue
                    else:
                        error_messages.append(f'Item {existing_item["itemName"]} has insufficient stock.')
                        continue
                else:
                    if existing_item['quantity'] >= sold_quantity:  
                        available_quantity = existing_item['quantity'] - sold_quantity
                        stockDate = existing_item['stockDate']
                    else:
                        error_messages.append(f'Item {existing_item["itemName"]} has insufficient stock.')
                        continue
                
                stock_id = existing_item['_id']
                if 'selling_price' in existing_item:
                    selling_price = existing_item['selling_price']
                    revenue = sold_quantity * existing_item['selling_price']
                else:
                    selling_price = 0
                    revenue = 0
                data = {
                    'itemName': existing_item['itemName'],
                    'quantity': sold_quantity,
                    'unitPrice': selling_price,
                    'saleDate': timestamp,
                    'company_name': company['company_name'],
                    'timestamp': timestamp,
                    'revenue': revenue,
                    'stockDate': stockDate,
                    'stock_id': stock_id
                }

                db.stock_sales.insert_one(data)
                db.audit_logs.insert_one({
                    'user': login_data,
                    'Activity': 'Added a new sale',
                    'Item': existing_item['itemName'],
                    'timestamp': datetime.now()
                })
                db.inventories.update_one({'company_name': company['company_name'], 'itemName': existing_item['itemName']}, {'$set': {'available_quantity': available_quantity}})
                success_messages.append(f'Sale for {existing_item["itemName"]} was successful')
            else:
                error_messages.append(f'Scanned item {product_id} does not exist')
        
        # Flash success messages if any
        if success_messages:
            flash(' | '.join(success_messages), 'success')
        
        # Flash error messages if any
        if error_messages:
            flash(' | '.join(error_messages), 'error')

        return redirect('/scan bar code page')
    else:
        flash('Scan Item list is empty','error')
        return redirect('/scan bar code page')
    
@stockManagement.route('/in-house-use', methods=['POST'])
def inhouse():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
                
            all_items = request.json.get('items', [])  # Access the JSON data sent from the client

            # Initialize empty lists for item details
            itemNames = []
            itemQuantities = []
            itemStockDates = []
            itemUnitPrices = []
            itemOldUnitPrices = []
            
            # Extract details from the first item
            if all_items:
                productName = all_items[0].get('productName', '')
                productQuantity = float(all_items[0].get('productQuantity', 0))
                productUnitOfMeasurement = all_items[0].get('unitOfMeasurement', '')
                productPrice = float(all_items[0].get('productPrice', 0))
                useDate = datetime.strptime(all_items[0].get('useDate', ''), '%Y-%m-%d')
                company_name = company.get('company_name', '')

                out_of_stock_items = []
                over_quantified = []
                in_stockID = []
                in_stockQty = []

                for item in all_items:
                    itemNames.append(item.get('itemName', ''))
                    itemQuantity = float(item.get('itemQuantity', 0))
                    itemQuantities.append(itemQuantity)

                    # Check if the item exists in the database
                    existing_item = db.inventories.find_one({
                        'itemName': item['itemName'],
                        'company_name': company_name
                    })

                    if existing_item:
                        if 'available_quantity' in existing_item:
                            if existing_item['available_quantity'] <= 0:
                                out_of_stock_items.append(item['itemName'])
                                flash(f'Item {item["itemName"]} is out of stock', 'error')
                                continue
                            if itemQuantity > existing_item['available_quantity']:
                                over_quantified.append(item['itemName'])
                                flash(f'Quantity for item {item["itemName"]} is too high', 'error')
                                continue
                            else:
                                available_quantity = existing_item['available_quantity'] - itemQuantity
                                itemStockDates.append(existing_item['stockDate'])
                                in_stockID.append(existing_item['_id'])
                                in_stockQty.append(available_quantity)
                                itemUnitPrices.append(existing_item['unitPrice'])
                                itemOldUnitPrices.append(existing_item.get('oldUnitPrice', 0))
                                flash(f'Inhouse use of {item["itemName"]} updated successfully', 'success')
                        else:
                            if itemQuantity > existing_item['quantity']:
                                over_quantified.append(item['itemName'])
                                flash(f'Quantity for item {item["itemName"]} is too high', 'error')
                            else:
                                available_quantity = existing_item['quantity'] - itemQuantity
                                itemStockDates.append(existing_item['stockDate'])
                                in_stockID.append(existing_item['_id'])
                                in_stockQty.append(available_quantity)
                                itemUnitPrices.append(existing_item['unitPrice'])
                                itemOldUnitPrices.append(existing_item.get('oldUnitPrice', 0))
                                flash(f'Inhouse use of {item["itemName"]} updated successfully', 'success')
                    else:
                        flash(f"Item {item['itemName']} does not exist.", 'error')

                if out_of_stock_items:
                    flash(f'The following items are out of stock: {", ".join(out_of_stock_items)}', 'error')
                if over_quantified:
                    flash(f'Please enter smaller quantities for the following items: {", ".join(over_quantified)}', 'error')

                if not out_of_stock_items and not over_quantified:
                    document = {
                        'productName': productName,
                        'productQuantity': productQuantity,
                        'productUnitOfMeasurement': productUnitOfMeasurement,
                        'productPrice': productPrice,
                        'useDate': useDate,
                        'itemName': itemNames,
                        'itemQuantity': itemQuantities,
                        'itemUnitPrices': itemUnitPrices,
                        'itemOldUnitPrices': itemOldUnitPrices,
                        'itemStockDates': itemStockDates,
                        'company_name': company_name
                    }
                    for id, available_quantity in zip(in_stockID, in_stockQty):
                        db.inventories.update_one({'_id': id}, {'$set': {'available_quantity': available_quantity}})
                    db.inhouse.insert_one(document)
                    db.audit_logs.insert_one({
                        'user': login_data,
                        'Activity': 'Inhouse production',
                        'Item': 'Items',
                        'timestamp': datetime.now()
                    })

            return jsonify({'redirect': url_for('stockManagement_route.update_production_activity')})
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/in-house-used-items', methods=['POST'])
def inhouse_used_items():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Please login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
                
            all_items = request.json.get('items', [])  # Access the JSON data sent from the client

            # Initialize lists for item details
            itemNames = []
            itemQuantities = []
            itemStockDates = []
            itemUseDates = []
            itemUnitPrices = []
            itemOldUnitPrices = []
            
            company_name = company.get('company_name', '')

            out_of_stock_items = []
            over_quantified = []
            in_stockID = []
            in_stockQty = []

            for item in all_items:
                itemName = item.get('usedItemName', '')
                itemQuantity = float(item.get('usedItemQuantity', 0))
                use_date = item.get('usedUseDate', '')
                useDate = datetime.strptime(use_date, '%Y-%m-%d') if use_date else None

                itemNames.append(itemName)
                itemQuantities.append(itemQuantity)
                itemUseDates.append(useDate)

                existing_item = db.inventories.find_one({
                    'itemName': itemName,
                    'company_name': company_name
                })

                if existing_item:
                    if 'available_quantity' in existing_item:
                        if existing_item['available_quantity'] <= 0:
                            out_of_stock_items.append(itemName)
                            flash(f'Item {itemName} is out of stock', 'error')
                            continue
                        if itemQuantity > existing_item['available_quantity']:
                            over_quantified.append(itemName)
                            flash(f'Quantity for item {itemName} is too high', 'error')
                            continue
                        available_quantity = existing_item['available_quantity'] - itemQuantity
                        itemStockDates.append(existing_item['stockDate'])
                        in_stockID.append(existing_item['_id'])
                        in_stockQty.append(available_quantity)
                        itemUnitPrices.append(existing_item['unitPrice'])
                        itemOldUnitPrices.append(existing_item.get('oldUnitPrice', 0))
                        flash(f'Inhouse use of {itemName} updated successfully', 'success')
                    else:
                        if itemQuantity > existing_item['quantity']:
                            over_quantified.append(itemName)
                            flash(f'Quantity for item {itemName} is too high', 'error')
                        else:
                            available_quantity = existing_item['quantity'] - itemQuantity
                            itemStockDates.append(existing_item['stockDate'])
                            in_stockID.append(existing_item['_id'])
                            in_stockQty.append(available_quantity)
                            itemUnitPrices.append(existing_item['unitPrice'])
                            itemOldUnitPrices.append(existing_item.get('oldUnitPrice', 0))
                            flash(f'Inhouse use of {itemName} updated successfully', 'success')
                else:
                    flash(f'Item {itemName} does not exist', 'error')

            if out_of_stock_items:
                flash(f'The following items are out of stock: {", ".join(out_of_stock_items)}', 'error')
            if over_quantified:
                flash(f'Please enter smaller quantities for the following items: {", ".join(over_quantified)}', 'error')

            if not out_of_stock_items and not over_quantified:
                document = {
                    'itemName': itemNames,
                    'itemQuantity': itemQuantities,
                    'itemUnitPrices': itemUnitPrices,
                    'itemOldUnitPrices': itemOldUnitPrices,
                    'itemStockDates': itemStockDates,
                    'useDate': itemUseDates,
                    'company_name': company_name
                }
                for id, available_quantity in zip(in_stockID, in_stockQty):
                    db.inventories.update_one({'_id': id}, {'$set': {'available_quantity': available_quantity}})
                db.inhouse_use.insert_one(document)
                db.audit_logs.insert_one({
                    'user': login_data,
                    'Activity': 'Inhouse use of items',
                    'Item': 'Items',
                    'timestamp': datetime.now()
                })

            return jsonify({'redirect': url_for('stockManagement_route.update_inhouse_use_page')})
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@stockManagement.route('/revenue-details')
def revenue_details():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            if session.get('view_revenue') == "no":
                flash("You do not have rights to view profits","error")
                return redirect('/stock-overview')
            else:
                company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
                
                subscription = db.managers.find_one({'name': company['company_name']}, {'account_type': 1, 'manager_email': 1, '_id': 0})
                account_type = subscription['account_type']
                # Remove any empty strings from the list
                account_type = [atype for atype in account_type if atype]

                if 'Enterprise Resource Planning' in account_type:
                    company_name = company['company_name']
                    now = datetime.now()
                    first_day_of_current_month = datetime(now.year, now.month, 1)
                    twelve_months_ago = first_day_of_current_month.replace(year=first_day_of_current_month.year - 1)

                    pipeline = [
                        {
                            '$match': {
                                'company_name': company_name,
                                'saleDate': {'$gte': twelve_months_ago}
                            }
                        },
                        {
                            '$group': {
                                '_id': {'itemName': '$itemName', 'stockDate': '$stockDate'},
                                'totalRevenue': {'$sum': '$revenue'},
                                'quantitySold': {'$sum': '$quantity'}
                            }
                        },
                        {
                            '$lookup': {
                                'from': 'inventories',
                                'let': {'itemName': '$_id.itemName', 'stockDate': '$_id.stockDate'},
                                'pipeline': [
                                    {
                                        '$match': {
                                            '$expr': {
                                                '$and': [
                                                    {'$eq': ['$itemName', '$$itemName']},
                                                    {'$eq': ['$company_name', company_name]},
                                                    { '$eq': ['$stockDate', '$$stockDate'] }
                                                ]
                                            }
                                        }
                                    },
                                    {
                                        '$project': {
                                            '_id': 0,
                                            'quantity': 1,
                                            'unitPrice': 1,
                                            'stockDate': 1
                                        }
                                    }
                                ],
                                'as': 'inventoryDetails'
                            }
                        },
                        {
                            '$lookup': {
                                'from': 'old_inventories',
                                'let': {'itemName': '$_id.itemName', 'stockDate': '$_id.stockDate'},
                                'pipeline': [
                                    {
                                        '$match': {
                                            '$expr': {
                                                '$and': [
                                                    {'$eq': ['$itemName', '$$itemName']},
                                                    {'$eq': ['$company_name', company_name]},
                                                    { '$eq': ['$stockDate', '$$stockDate'] }
                                                ]
                                            }
                                        }
                                    },
                                    {
                                        '$project': {
                                            '_id': 0,
                                            'quantity': 1,
                                            'unitPrice': 1,
                                            'stockDate': 1
                                        }
                                    }
                                ],
                                'as': 'oldInventoryDetails'
                            }
                        },
                        {
                            '$project': {
                                'inventoryDetails': {
                                    '$cond': {
                                        'if': {'$gt': [{'$size': '$inventoryDetails'}, 0]},
                                        'then': '$inventoryDetails',
                                        'else': '$oldInventoryDetails'
                                    }
                                },
                                'totalRevenue': 1,
                                'quantitySold': 1,
                                '_id': '$_id'
                            }
                        },
                        {
                            '$unwind': '$inventoryDetails'
                        },
                        {
                            '$group': {
                                '_id': '$_id.itemName',
                                'stockDate': {'$first': '$_id.stockDate'},
                                'totalRevenue': {'$first': '$totalRevenue'},
                                'quantitySold': {'$first': '$quantitySold'},
                                'unitPrice': {'$avg': '$inventoryDetails.unitPrice'}
                            }
                        },
                        {
                            '$addFields': {
                                'unitProfit': {
                                    '$round': [
                                        {
                                            '$subtract': [
                                                {'$divide': ['$totalRevenue', '$quantitySold']},
                                                '$unitPrice'
                                            ]
                                        },
                                        0
                                    ]
                                },
                                'totalProfit': {
                                    '$subtract': [
                                        '$totalRevenue',
                                        {'$multiply': ['$unitPrice', '$quantitySold']}
                                    ]
                                }
                            }
                        },
                        {
                            '$match': {
                                'quantitySold': {'$ne': 0}
                            }
                        }
                    ]

                    revenue_info = list(db.stock_sales.aggregate(pipeline))
                    revenue_info.sort(key=lambda x: x['_id'])
                    dp = company.get('dp')
                    dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
                    return render_template('revenue info.html', revenue_info = revenue_info, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/sales-details')
def sales_details():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            if session.get('view_sales') == "no":
                flash("You do not have rights to view sales","error")
                return redirect('/stock-overview')
            else:
                company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})

                company_name = company['company_name']

                now = datetime.now()
                first_day_of_current_month = datetime(now.year, now.month, 1)
                twelve_months_ago = first_day_of_current_month.replace(year=first_day_of_current_month.year - 1)

                sales_info = list(db.stock_sales.find({'company_name': company_name, 'saleDate': {'$gte': twelve_months_ago}}))
                sales_info.sort(key=lambda x: x['saleDate'], reverse=True)
                dp = company.get('dp')
                dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
                return render_template('sales info.html', sales_info = sales_info, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/stock-details')
def stock_details():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})
            
            subscription = db.managers.find_one({'name': company['company_name']}, {'account_type': 1, 'manager_email': 1, '_id': 0})
            account_type = subscription['account_type']
            # Remove any empty strings from the list
            account_type = [atype for atype in account_type if atype]

            if 'Enterprise Resource Planning' in account_type:
                company_name = company['company_name']
                stock_info = list(db.inventories.find({'company_name': company_name}))
                stock_info.sort(key=lambda x: x.get('timestamp', x['stockDate']), reverse=True)
                stock_info.sort(key=lambda x: x['itemName'])

                dp = company.get('dp')
                dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
                return render_template('stock info.html', stock_info = stock_info, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
        
@stockManagement.route('/stock-history-details')
def stock_history_details():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})
            
            subscription = db.managers.find_one({'name': company['company_name']}, {'account_type': 1, 'manager_email': 1, '_id': 0})
            account_type = subscription['account_type']
            # Remove any empty strings from the list
            account_type = [atype for atype in account_type if atype]

            if 'Enterprise Resource Planning' in account_type:
                company_name = company['company_name']
                now = datetime.now()
                first_day_of_current_month = datetime(now.year, now.month, 1)
                twelve_months_ago = first_day_of_current_month.replace(year=first_day_of_current_month.year - 1)
                stock_info = list(db.old_inventories.find({'company_name': company_name, 'stockDate': {'$gte': twelve_months_ago}}))
                stock_info.sort(key=lambda x: x['stockDate'], reverse=True)

                dp = company.get('dp')
                dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
                return render_template('stock history.html', stock_info = stock_info, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/inhouse-item-use-details')
def inhouse_items_use_details():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})
            
            company_name = company['company_name']
            now = datetime.now()
            first_day_of_current_month = datetime(now.year, now.month, 1)
            twelve_months_ago = first_day_of_current_month.replace(year=first_day_of_current_month.year - 1)
            inhouse_item_use = list(db.inhouse_use.find({'company_name': company_name, 'useDate': {'$gte': twelve_months_ago}}))
            inhouse_item_use.sort(key=lambda x: max(x['useDate']), reverse=True)

            available_itemNames = []
            items_to_update = []
            available_items = list(db.inventories.find({'company_name': company['company_name']}))
            if len(available_items) != 0:
                for item in available_items:
                    if 'available_quantity' in item:
                        if item['available_quantity'] > 0:
                            available_itemNames.append(item['itemName'])
                    else:
                        available_itemNames.append(item['itemName'])
                for item in available_items:
                    items_to_update.append(item['itemName'])
            dp = company.get('dp')
            dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
            return render_template('inhouse item use info.html', inhouse_item_use = inhouse_item_use, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/stock-overview', methods=["GET", "POST"])
def stock_overview():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})

            company_name = company['company_name']
            
            if session.get('is_manager') != "is_manager":
                flash("You do not have rights to view the dashboard","error")
                dp = company.get('dp')
                dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
                return render_template('stock dashboard.html', 
                       profits_chart=None,
                       Losses_chart=None,
                       revenue=None,
                       quantity_sold_stocked=None,
                       trended_profit=None,
                       start_of_previous_month=None,
                       first_day_of_current_month=None,dp=dp_str)
            else:
                session.pop("profits_chart", None)
                session.pop("loss_chart", None)
                session.pop("revenue_and_qty_chart", None)
                session.pop("monthly_profits_chart", None)
                session.pop("inhouse_costs_chart", None)
                session.pop("inhouse_revenue_chart", None)

                startdate_on_str = request.form.get("startdate")
                enddate_on_str = request.form.get("enddate")

                if startdate_on_str and enddate_on_str:
                    start_of_previous_month = datetime.strptime(startdate_on_str, '%Y-%m-%d')
                    first_day_of_current_month = datetime.strptime(enddate_on_str, '%Y-%m-%d')
                else:
                    today = datetime.today()

                    start_of_previous_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                    first_day_of_current_month = today.replace(hour=0, minute=0, second=0, microsecond=0)

                pipeline = [
                    {
                        '$match': {
                            'company_name': company_name,
                            'saleDate': {
                                '$gte': start_of_previous_month,
                                '$lte': first_day_of_current_month
                            }
                        }
                    },
                    {
                        '$group': {
                            '_id': {'itemName': '$itemName', 'stockDate': '$stockDate'},
                            'totalRevenue': {'$sum': '$revenue'},
                            'quantitysold': {'$sum': '$quantity'}
                        }
                    },
                    {
                        '$lookup': {
                            'from': 'inventories',
                            'let': {'itemName': '$_id.itemName', 'stockDate': '$_id.stockDate'},
                            'pipeline': [
                                {
                                    '$match': {
                                        '$expr': {
                                            '$and': [
                                                {'$eq': ['$itemName', '$$itemName']},
                                                {'$eq': ['$company_name', company_name]},
                                                { '$eq': ['$stockDate', '$$stockDate'] }
                                            ]
                                        }
                                    }
                                },
                                {
                                    '$project': {
                                        '_id': 0,
                                        'quantity': 1,
                                        'unitPrice': 1,
                                        'stockDate': 1,
                                        'totalPrice': 1
                                    }
                                }
                            ],
                            'as': 'inventoryDetails'
                        }
                    },
                    {
                        '$lookup': {
                            'from': 'old_inventories',
                            'let': {'itemName': '$_id.itemName', 'stockDate': '$_id.stockDate'},
                            'pipeline': [
                                {
                                    '$match': {
                                        '$expr': {
                                            '$and': [
                                                {'$eq': ['$itemName', '$$itemName']},
                                                {'$eq': ['$company_name', company_name]},
                                                { '$eq': ['$stockDate', '$$stockDate'] }
                                            ]
                                        }
                                    }
                                },
                                {
                                    '$project': {
                                        '_id': 0,
                                        'quantity': 1,
                                        'unitPrice': 1,
                                        'stockDate': 1,
                                        'totalPrice': 1
                                    }
                                }
                            ],
                            'as': 'oldInventoryDetails'
                        }
                    },
                    {
                        '$project': {
                            'inventoryDetails': {
                                '$cond': {
                                    'if': {'$gt': [{'$size': '$inventoryDetails'}, 0]},
                                    'then': '$inventoryDetails',
                                    'else': '$oldInventoryDetails'
                                }
                            },
                            'totalRevenue': 1,
                            'quantitysold': 1,
                            '_id': 1
                        }
                    }
                ]

                revenue_info = list(db.stock_sales.aggregate(pipeline))
                item_names = []
                quantities_sold = []
                quantities_stocked = []
                total_revenues = []
                total_prices = []
                profits = []

                for record in revenue_info:
                    item_names.append(record['_id']['itemName'])
                    quantities_sold.append(record['quantitysold'])
                    total_revenues.append(record['totalRevenue'])

                    # Initialize variables for each iteration
                    quantities_stocked_iter = 0
                    total_price_iter = 0
                    profit_iter = 0

                    # Check if 'inventoryDetails' is in record and has the necessary structure
                    if 'inventoryDetails' in record and record['inventoryDetails']:
                        quantity_stocked = record['inventoryDetails'][0].get('quantity', 0)
                        total_price_iter = (record['inventoryDetails'][0].get('unitPrice', 0))*record['quantitysold']
                        profit_iter = record['totalRevenue'] - total_price_iter

                    # Append values for this iteration to the lists
                    quantities_stocked.append(quantities_stocked_iter)
                    total_prices.append(total_price_iter)
                    profits.append(profit_iter)

                # Create the DataFrame
                df_ungrouped = pd.DataFrame({
                    'Item Name': item_names,
                    'Quantity Sold': quantities_sold,
                    'Quantity Stocked': quantities_stocked,
                    'Total Revenue': total_revenues,
                    'Total Price': total_prices,
                    'Profit': profits
                })

                # Group by 'Item Name' and aggregate using sum
                df = df_ungrouped.groupby('Item Name', as_index=False).agg({
                    'Quantity Sold': 'sum',
                    'Quantity Stocked': 'sum',
                    'Total Revenue': 'sum',
                    'Total Price': 'sum',
                    'Profit': 'sum'
                })

                #####PLOTS
                #profits and losses
                # Filter positive profits
                profitableItems = []
                profits = []
                top_profitable_items = df[df['Profit'] > 0].sort_values(by='Profit', ascending=False).head(10)

                if not top_profitable_items.empty:
                    session['profits_chart'] = 'profits_chart'
                    for index, row in top_profitable_items.iterrows():
                        profitableItems.append(row['Item Name'])
                        profits.append(row['Profit'])
                top10profits = list(zip(profitableItems, profits))

                # Filter negative profits
                unprofitableItems = []
                losses = []
                top_unprofitable_items = df[df['Profit'] < 0].sort_values(by='Profit', ascending=False).head(10)

                if not top_unprofitable_items.empty:
                    session['loss_chart'] = 'loss_chart'
                    for index, row in top_unprofitable_items.iterrows():
                        unprofitableItems.append(row['Item Name'])
                        losses.append(row['Profit'])
                top10losses = list(zip(unprofitableItems, losses))

                ##total revenue
                revenueItems = []
                revenue = []
                if not df.empty:
                    session['revenue_and_qty_chart'] = 'revenue_and_qty_chart'
                    top_revenue = df.sort_values(by='Total Revenue', ascending=False).head(10)
                    for index, row in top_revenue.iterrows():
                        revenueItems.append(row['Item Name'])
                        revenue.append(row['Total Revenue'])
                top10revenues = list(zip(revenueItems, revenue))

                ##Quantity sold
                soldItems = []
                soldQuantity = []
                if not df.empty:
                    quantity_sold = df.sort_values(by='Quantity Sold', ascending=False).head(10)
                    for index, row in quantity_sold.iterrows():
                        soldItems.append(row['Item Name'])
                        soldQuantity.append(row['Quantity Sold'])
                top10SoldItems = list(zip(soldItems, soldQuantity))

                ###PROFIT TRENDS
                now = datetime.now()
                first_day_of_current_month = datetime(now.year, now.month, 1)
                twelve_months_ago = first_day_of_current_month.replace(year=first_day_of_current_month.year - 1)
                pipeline_profits = [
                    {
                        '$match': {
                            'company_name': company_name,
                            'saleDate': {'$gte': twelve_months_ago}
                        }
                    },
                    {
                        '$group': {
                            '_id': {'itemName': '$itemName', 'stockDate': '$stockDate'},
                            'totalRevenue': {'$sum': '$revenue'},
                            'quantitySold': {'$sum': '$quantity'}
                        }
                    },
                    {
                        '$lookup': {
                            'from': 'inventories',
                            'let': {'itemName': '$_id.itemName', 'stockDate': '$_id.stockDate'},
                            'pipeline': [
                                {
                                    '$match': {
                                        '$expr': {
                                            '$and': [
                                                {'$eq': ['$itemName', '$$itemName']},
                                                {'$eq': ['$company_name', company_name]},
                                                { '$eq': ['$stockDate', '$$stockDate'] }
                                            ]
                                        }
                                    }
                                },
                                {
                                    '$project': {
                                        '_id': 0,
                                        'quantity': 1,
                                        'unitPrice': 1,
                                        'stockDate': 1
                                    }
                                }
                            ],
                            'as': 'inventoryDetails'
                        }
                    },
                    {
                        '$lookup': {
                            'from': 'old_inventories',
                            'let': {'itemName': '$_id.itemName', 'stockDate': '$_id.stockDate'},
                            'pipeline': [
                                {
                                    '$match': {
                                        '$expr': {
                                            '$and': [
                                                {'$eq': ['$itemName', '$$itemName']},
                                                {'$eq': ['$company_name', company_name]},
                                                { '$eq': ['$stockDate', '$$stockDate'] }
                                            ]
                                        }
                                    }
                                },
                                {
                                    '$project': {
                                        '_id': 0,
                                        'quantity': 1,
                                        'unitPrice': 1,
                                        'stockDate': 1
                                    }
                                }
                            ],
                            'as': 'oldInventoryDetails'
                        }
                    },
                    {
                        '$project': {
                            'inventoryDetails': {
                                '$cond': {
                                    'if': {'$gt': [{'$size': '$inventoryDetails'}, 0]},
                                    'then': '$inventoryDetails',
                                    'else': '$oldInventoryDetails'
                                }
                            },
                            'totalRevenue': 1,
                            'quantitySold': 1,
                            '_id': '$_id'
                        }
                    },
                    {
                        '$unwind': '$inventoryDetails'
                    },
                    {
                        '$group': {
                            '_id': '$_id.itemName',
                            'stockDate': {'$first': '$_id.stockDate'},
                            'totalRevenue': {'$first': '$totalRevenue'},
                            'quantitySold': {'$first': '$quantitySold'},
                            'unitPrice': {'$avg': '$inventoryDetails.unitPrice'}
                        }
                    },
                    {
                        '$addFields': {
                            'unitProfit': {
                                '$round': [
                                    {
                                        '$subtract': [
                                            {'$divide': ['$totalRevenue', '$quantitySold']},
                                            '$unitPrice'
                                        ]
                                    },
                                    0
                                ]
                            },
                            'totalProfit': {
                                '$subtract': [
                                    '$totalRevenue',
                                    {'$multiply': ['$unitPrice', '$quantitySold']}
                                ]
                            }
                        }
                    },
                    {
                        '$match': {
                            'quantitySold': {'$ne': 0}
                        }
                    }
                ]
                
                profit_info = list(db.stock_sales.aggregate(pipeline_profits))

                profit_item_names = []
                profit_data = []
                profit_stock_dates = []


                for profit_record in profit_info:
                    profit_item_names.append(profit_record['_id'])
                    profit_stock_dates.append(profit_record['stockDate'])
                    profit_data.append(profit_record['totalProfit'])

                # Create the DataFrame
                profit_info_df = pd.DataFrame({
                    'Item Name': profit_item_names,
                    'Profit': profit_data,
                    'Stock Date': profit_stock_dates
                })

                # Convert 'Stock Date' to datetime
                profit_info_df['Stock Date'] = pd.to_datetime(profit_info_df['Stock Date'])

                # Group by month and calculate the sum of profits
                monthly_profits = profit_info_df.groupby(profit_info_df['Stock Date'].dt.to_period('M'))['Profit'].sum()

                # Create a DataFrame with month names
                monthly_profits_df = pd.DataFrame({
                    'Month': monthly_profits.index.to_timestamp().strftime('%B'),
                    'Monthly Profit': monthly_profits
                })

                # Create the line chart
                if not monthly_profits_df.empty:
                    session['monthly_profits_chart'] = 'monthly_profits_chart'

                trended_profit = {
                    'labels': monthly_profits_df['Month'].tolist(),
                    'values': monthly_profits_df['Monthly Profit'].tolist()
                }

                del df_ungrouped, df, profit_info_df, monthly_profits, monthly_profits_df
                gc.collect()
                dp = company.get('dp')
                dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
                return render_template('stock dashboard.html',top10profits=top10profits,top10losses=top10losses,top10revenues=top10revenues,
                                    top10SoldItems=top10SoldItems,trended_profit=trended_profit,
                                    start_of_previous_month=start_of_previous_month,
                                    first_day_of_current_month=first_day_of_current_month, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

###DOANLOAD STOCK DATA   
@stockManagement.route('/download-stock-data', methods=["POST"])
def download_stock_data():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            startdate_on_str = request.form.get("startdate")
            enddate_on_str = request.form.get("enddate")
            startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
            enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')

            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})

            current_stock = db.inventories.find(
                {'company_name': company['company_name'], 'stockDate': {'$gte': startdate, '$lte': enddate}},
                {'_id': 0, 'company_name': 0}
            )
            old_stock = db.old_inventories.find(
                {'company_name': company['company_name'], 'stockDate': {'$gte': startdate, '$lte': enddate}},
                {'_id': 0, 'company_name': 0}
            )
            combined_stock = list(current_stock) + list(old_stock)

            # Sort data by stockDate in descending order
            sorted_combined_stock = sorted(combined_stock, key=lambda x: x["stockDate"], reverse=True)

            # Create Excel file
            excel_buffer = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Data"

            # Write header row
            headers = ['Item Name', 'Stock Date', 'Stocked Quantity', 'Available Stock', 'Measurement', 'Buying Price', 'Total Buying Price', 'Total Stock Value']
            ws.append(headers)

            # Write data rows
            for record in sorted_combined_stock:
                if 'cumulativeOldPrices' in record:
                    cumulativeOldPrices = record['cumulativeOldPrices']
                elif 'oldTotalPrice' in record:
                    cumulativeOldPrices = record['oldTotalPrice'] + record['totalPrice']
                else:
                    cumulativeOldPrices = record['totalPrice']
                row = [
                    record.get('itemName', ''),
                    record.get('stockDate', '').strftime('%Y-%m-%d') if isinstance(record.get('stockDate'), datetime) else '',
                    record.get('quantity', 0),
                    record.get('available_quantity', 0),
                    record.get('unitOfMeasurement', ''),
                    record.get('unitPrice', 0),
                    record.get('totalPrice', 0),
                    cumulativeOldPrices
                ]
                ws.append(row)

            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Create the response
            response = make_response(excel_buffer.getvalue())
            response.headers['Content-Disposition'] = f"attachment; filename={company['company_name']}_Stock_Data_{startdate_on_str}_{enddate_on_str}.xlsx"
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            # Clean up
            del wb
            del excel_buffer
            gc.collect()

            return response
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
###DOANLOAD REVENUE DATA   
@stockManagement.route('/download-revenue-data', methods=["POST"])
def download_revenue_data():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            startdate_on_str = request.form.get("startdate")
            enddate_on_str = request.form.get("enddate")
            startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
            enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})
            

            company_name = company['company_name']

            pipeline = [
                {
                    '$match': {
                        'company_name': company_name,
                        'saleDate': {
                            '$gte': startdate,
                            '$lte': enddate
                        }
                    }
                },
                {
                    '$group': {
                        '_id': {'itemName': '$itemName', 'stockDate': '$stockDate'},
                        'totalRevenue': {'$sum': '$revenue'},
                        'quantitySold': {'$sum': '$quantity'}
                    }
                },
                {
                    '$lookup': {
                        'from': 'inventories',
                        'let': {'itemName': '$_id.itemName', 'stockDate': '$_id.stockDate'},
                        'pipeline': [
                            {
                                '$match': {
                                    '$expr': {
                                        '$and': [
                                            {'$eq': ['$itemName', '$$itemName']},
                                            {'$eq': ['$company_name', company_name]},
                                            { '$eq': ['$stockDate', '$$stockDate'] }
                                        ]
                                    }
                                }
                            },
                            {
                                '$project': {
                                    '_id': 0,
                                    'quantity': 1,
                                    'unitPrice': 1,
                                    'stockDate': 1
                                }
                            }
                        ],
                        'as': 'inventoryDetails'
                    }
                },
                {
                    '$lookup': {
                        'from': 'old_inventories',
                        'let': {'itemName': '$_id.itemName', 'stockDate': '$_id.stockDate'},
                        'pipeline': [
                            {
                                '$match': {
                                    '$expr': {
                                        '$and': [
                                            {'$eq': ['$itemName', '$$itemName']},
                                            {'$eq': ['$company_name', company_name]},
                                            { '$eq': ['$stockDate', '$$stockDate'] }
                                        ]
                                    }
                                }
                            },
                            {
                                '$project': {
                                    '_id': 0,
                                    'quantity': 1,
                                    'unitPrice': 1,
                                    'stockDate': 1
                                }
                            }
                        ],
                        'as': 'oldInventoryDetails'
                    }
                },
                {
                    '$project': {
                        'inventoryDetails': {
                            '$cond': {
                                'if': {'$gt': [{'$size': '$inventoryDetails'}, 0]},
                                'then': '$inventoryDetails',
                                'else': '$oldInventoryDetails'
                            }
                        },
                        'totalRevenue': 1,
                        'quantitySold': 1,
                        '_id': '$_id'
                    }
                },
                {
                    '$unwind': '$inventoryDetails'
                },
                {
                    '$group': {
                        '_id': '$_id.itemName',
                        'stockDate': {'$first': '$_id.stockDate'},
                        'totalRevenue': {'$first': '$totalRevenue'},
                        'quantitySold': {'$first': '$quantitySold'},
                        'unitPrice': {'$avg': '$inventoryDetails.unitPrice'}  # Assuming you want the average unit price
                    }
                },
                {
                    '$addFields': {
                        'unitProfit': {
                            '$round': [
                                {
                                    '$subtract': [
                                        {'$divide': ['$totalRevenue', '$quantitySold']},
                                        '$unitPrice'
                                    ]
                                },
                                0
                            ]
                        },
                        'totalProfit': {
                            '$subtract': [
                                '$totalRevenue',
                                {'$multiply': ['$unitPrice', '$quantitySold']}
                            ]
                        }
                    }
                },
                {
                    '$match': {
                        'quantitySold': {'$ne': 0}
                    }
                }
            ]

            revenue_info = list(db.stock_sales.aggregate(pipeline))
            revenue_info.sort(key=lambda x: x['_id'])

            data_rows = []

            if revenue_info:
                for revenue in revenue_info:
                    item_name = revenue['_id']
                    stock_date = revenue['stockDate'].strftime('%Y-%m-%d')
                    buying_price = revenue['unitPrice']
                    quantity_sold = revenue['quantitySold']
                    total_revenue = revenue['totalRevenue']                   

                    unitProfit = round(revenue['unitProfit'],0)
                    totalProfit = revenue['totalProfit']

                    data_rows.append({
                        'Item Name': item_name,
                        'Stock Date': stock_date,
                        'Buying Price': buying_price,
                        'Sold Quantity': quantity_sold,
                        'Total Selling Price': total_revenue,
                        'Average Unit Profit': unitProfit,
                        'Total Profit': totalProfit
                    })

            df = pd.DataFrame(data_rows)

            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='Profits Data', index=False)
            excel_buffer.seek(0)

            del df, revenue_info, data_rows
            gc.collect()
            # Create the response
            response = make_response(excel_buffer.getvalue())
            response.headers['Content-Disposition'] = f"attachment; filename={company['company_name']}_Profits_Data_{startdate_on_str}_{enddate_on_str}.xlsx"
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            return response
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
###DOANLOAD SALES DATA   
@stockManagement.route('/download-sales-data', methods=["POST"])
def download_sales_data():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            startdate_on_str = request.form.get("startdate")
            enddate_on_str = request.form.get("enddate")
            startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
            enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})

            company_name = company['company_name']

            sales_info = list(db.stock_sales.find({
                'company_name': company_name,
                'saleDate': {'$gte': startdate, '$lte': enddate}
            }, {
                '_id': 0,
                'company_name': 0,
                'stockDate': 0
            }))

            # Sort sales info by saleDate in descending order
            sorted_sales_info = sorted(sales_info, key=lambda x: x["saleDate"], reverse=True)

            # Create Excel file
            excel_buffer = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Data"

            # Write header row
            headers = ['Item Name', 'Sale Date', 'Sold Quantity', 'Selling Price', 'Total Selling Price']
            ws.append(headers)

            # Write data rows
            for sale in sorted_sales_info:
                row = [
                    sale.get('itemName', ''),
                    sale.get('saleDate', '').strftime('%Y-%m-%d') if isinstance(sale.get('saleDate'), datetime) else '',
                    sale.get('quantity', 0),
                    sale.get('unitPrice', 0),
                    sale.get('revenue', 0),
                ]
                ws.append(row)

            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Create the response
            response = make_response(excel_buffer.getvalue())
            response.headers['Content-Disposition'] = f"attachment; filename={company['company_name']}_Sales_Data_{startdate_on_str}_{enddate_on_str}.xlsx"
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            # Clean up
            del wb
            del excel_buffer
            gc.collect()

            return response
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

# Function to calculate total production cost
def calculate_total_cost(row):
    total_cost = 0
    for qty, prices in zip(row['Item Quantity'], row['Item Unit Price']):
        total_cost += np.sum(np.array(qty) * np.array(prices))
    return total_cost

@stockManagement.route('/view-production-info')
def view_production_info():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            now = datetime.now()
            first_day_of_current_month = datetime(now.year, now.month, 1)
            twelve_months_ago = first_day_of_current_month.replace(year=first_day_of_current_month.year - 1)
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})

            company_name = company['company_name']

            inhouse_info = list(db.inhouse.find({'company_name': company_name, 'useDate': {'$gte': twelve_months_ago}}, {'company_name': 0}))

            if not inhouse_info:
                flash('No inhouse production data available for the past 12 months.', 'info')
                return render_template('production info.html', inhouse_df=None, dp=None)

            inhouse_product_ids = []
            inhouse_productName = []
            inhouse_productUnitOfMeasurement = []
            inhouse_productQuantity = []
            inhouse_productPrice = []
            inhouse_useDate = []
            inhouse_itemName = []
            inhouse_itemQuantity = []
            inhouse_itemUnitPrices = []
            inhouse_itemStockDates = []

            for record in inhouse_info:
                productID = record['_id']
                productName = record['productName']
                productQuantity = record['productQuantity']
                productMeasure = record['productUnitOfMeasurement']
                productPrice = record['productPrice']
                useDate = record['useDate']
                item_name = record['itemName']
                item_quantity = record['itemQuantity']
                item_unit_price = record['itemUnitPrices']
                itemStockDates = record['itemStockDates']
            
                inhouse_product_ids.append(productID)
                inhouse_productName.append(productName)
                inhouse_productQuantity.append(productQuantity)
                inhouse_productUnitOfMeasurement.append(productMeasure)
                inhouse_productPrice.append(productPrice)
                inhouse_useDate.append(useDate)
                inhouse_itemName.append(item_name)
                inhouse_itemQuantity.append(item_quantity)
                inhouse_itemUnitPrices.append(item_unit_price)
                inhouse_itemStockDates.append(itemStockDates)

            # Create the DataFrame
            inhouse_df = pd.DataFrame({
                'Product ID': inhouse_product_ids,
                'Product Name': inhouse_productName,
                'Product Quantity': inhouse_productQuantity,
                'Product Measurement': inhouse_productUnitOfMeasurement,
                'Product Unit Price': inhouse_productPrice,
                'Date Produced': inhouse_useDate,
                'Item Used': inhouse_itemName,
                'Item Quantity': inhouse_itemQuantity,
                'Item Unit Price': inhouse_itemUnitPrices,
                'Item Stock Date': inhouse_itemStockDates
            })

            # Define total production cost calculation
            def calculate_total_cost(row):
                total_cost = sum(qty * price for qty, price in zip(row['Item Quantity'], row['Item Unit Price']))
                return total_cost

            # Apply the function to each row to calculate 'Total Production Cost'
            inhouse_df['Total Production Cost'] = inhouse_df.apply(calculate_total_cost, axis=1)
            inhouse_df_sorted = inhouse_df.sort_values(by='Date Produced')
            dp = company.get('dp')
            dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
            return render_template('production info.html', inhouse_df=inhouse_df_sorted, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
       
###DOANLOAD SALES DATA   
@stockManagement.route('/download-inhouse-data', methods=["POST"])
def download_inhouse():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            startdate_on_str = request.form.get("startdate")
            enddate_on_str = request.form.get("enddate")
            startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
            enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})

            company_name = company['company_name']

            inhouse_info = list(db.inhouse.find({
                'company_name': company_name,
                'useDate': {'$gte': startdate, '$lte': enddate}
            }, {
                'company_name': 0
            }))

            inhouse_product_ids = []
            inhouse_productName = []
            inhouse_productQuantity = []
            inhouse_productPrice = []
            inhouse_useDate = []
            inhouse_itemName = []
            inhouse_itemQuantity = []
            inhouse_itemUnitPrices = []
            inhouse_itemStockDates = []

            for record in inhouse_info:
                productID = record['_id']
                productName = record['productName']
                productQuantity = record['productQuantity']
                productPrice = record['productPrice']
                useDate = record['useDate']
                item_name = record['itemName']
                item_quantity = record['itemQuantity']
                item_unit_price = record['itemUnitPrices']
                itemStockDates = record['itemStockDates']
            
                inhouse_product_ids.append(productID)
                inhouse_productName.append(productName)
                inhouse_productQuantity.append(productQuantity)
                inhouse_productPrice.append(productPrice)
                inhouse_useDate.append(useDate)
                inhouse_itemName.append(item_name)
                inhouse_itemQuantity.append(item_quantity)
                inhouse_itemUnitPrices.append(item_unit_price)
                inhouse_itemStockDates.append(itemStockDates)

            # Create the DataFrame
            inhouse_df = pd.DataFrame({
                'Product ID': inhouse_product_ids,
                'Product Name': inhouse_productName,
                'Product Quantity': inhouse_productQuantity,
                'Product Unit Price': inhouse_productPrice,
                'Date Produced': inhouse_useDate,
                'Item Used': inhouse_itemName,
                'Item Quantity': inhouse_itemQuantity,
                'Item Unit Price': inhouse_itemUnitPrices,
                'Item Stock Date': inhouse_itemStockDates
            })

            # Apply the function to each row to calculate 'Total Production Cost'
            inhouse_df['Total Production Cost'] = inhouse_df.apply(calculate_total_cost, axis=1)
            inhouse_df_sorted = inhouse_df.sort_values(by='Date Produced')

            # Create an in-memory buffer for the Excel file
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                inhouse_df_sorted.to_excel(writer, sheet_name='Inhouse Data', index=False)
            excel_buffer.seek(0)

            # Create a zip file containing the Excel file
            zip_buffer = BytesIO()
            with ZipFile(zip_buffer, 'w') as zip_file:
                zip_file.writestr(f"{company['company_name']}_Inhouse_Data_{startdate_on_str}_{enddate_on_str}.xlsx", excel_buffer.read())
            
            zip_buffer.seek(0)
            zip_data = zip_buffer.getvalue()
            del inhouse_df, inhouse_df_sorted
            gc.collect()
            # Create the response
            response = make_response(zip_data)
            response.headers['Content-Disposition'] = f"attachment; filename={company['company_name']}_Inhouse_Data_{startdate_on_str}_{enddate_on_str}.zip"
            response.headers['Content-Type'] = 'application/zip'

            return response
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@stockManagement.route('/download-inhouse-item-data', methods=["POST"])
def download_inhouse_item_use():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            startdate_on_str = request.form.get("startdate")
            enddate_on_str = request.form.get("enddate")
            startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
            enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})

            company_name = company['company_name']

            inhouse_info = list(db.inhouse_use.find({
                'company_name': company_name,
                'useDate': {'$gte': startdate, '$lte': enddate}
            }, {
                'company_name': 0
            }))

            inhouse_itemName = []
            inhouse_useDate = []
            inhouse_itemQuantity = []
            inhouse_itemAverageUnitPrices = []
            inhouse_itemStockDates = []

            for record in inhouse_info:
                item_name = record['itemName']
                useDate = record['useDate']
                item_quantity = record['itemQuantity']
                if 'oldUnitPrice' in record:
                    average_unit_price = (record['oldUnitPrice'] + record['itemUnitPrices']) / 2
                else:
                    average_unit_price = record['itemUnitPrices']
                itemStockDates = record['itemStockDates']
            
                inhouse_itemName.append(item_name)
                inhouse_useDate.append(useDate)
                inhouse_itemQuantity.append(item_quantity)
                inhouse_itemAverageUnitPrices.append(average_unit_price)
                inhouse_itemStockDates.append(itemStockDates)

            # Create the DataFrame
            inhouse_df = pd.DataFrame({
                'Item Used': inhouse_itemName,
                'Date Used': inhouse_useDate,
                'Item Stock Date': inhouse_itemStockDates,
                'Item Quantity': inhouse_itemQuantity,
                'Item Average Price': inhouse_itemAverageUnitPrices,
            })

            # Explode DataFrame to handle lists in columns
            inhouse_df_exploded = inhouse_df.explode('Item Used')
            inhouse_df_exploded['Date Used'] = inhouse_df.explode('Date Used')['Date Used']
            inhouse_df_exploded['Item Stock Date'] = inhouse_df.explode('Item Stock Date')['Item Stock Date']
            inhouse_df_exploded['Item Quantity'] = inhouse_df.explode('Item Quantity')['Item Quantity']
            inhouse_df_exploded['Item Average Price'] = inhouse_df.explode('Item Average Price')['Item Average Price']
            inhouse_df_exploded.reset_index(drop=True, inplace=True)  # Reset the index

            inhouse_df_exploded['Average Total Cost'] = inhouse_df_exploded['Item Quantity'] * inhouse_df_exploded['Item Average Price']

            # Create an in-memory buffer for the Excel file
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                inhouse_df_exploded.to_excel(writer, sheet_name='Inhouse item use data', index=False)
            excel_buffer.seek(0)

            # Create a zip file containing the Excel file
            zip_buffer = BytesIO()
            with ZipFile(zip_buffer, 'w') as zip_file:
                zip_file.writestr(f"{company['company_name']}_Inhouse_Item_Use_Data_{startdate_on_str}_{enddate_on_str}.xlsx", excel_buffer.read())
            
            zip_buffer.seek(0)
            zip_data = zip_buffer.getvalue()
            del inhouse_df, inhouse_df_exploded
            gc.collect()
            # Create the response
            response = make_response(zip_data)
            response.headers['Content-Disposition'] = f"attachment; filename={company['company_name']}_Inhouse_Item_Use_Data_{startdate_on_str}_{enddate_on_str}.zip"
            response.headers['Content-Type'] = 'application/zip'

            return response
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

######add expenses
@stockManagement.route('/expenses-page')
def expenses_page():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            
            if company.get('update_sales') in ('yes', None):
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                return render_template('stock expenses.html', dp=dp_str)
            else:
                flash('You do not have rights to add expenses', 'error')
                return redirect('/stock-details')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/add-new-expense', methods=['POST'])
def add_new_expense():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
                
            all_items = request.json.get('items', [])  # Access the JSON data sent from the client
            timestamp = datetime.now()

            for expense in all_items:
                expense['expenseName'] = expense.get('expenseName', '').strip()
                
                try:
                    # Convert 'quantity' and 'unitPrice' to floats
                    expense['amount'] = float(expense.get('amount', 0))
                    expense['expenseDate'] = datetime.strptime(expense.get('expenseDate', ''), '%Y-%m-%d')

                    expense['company_name'] = company.get('company_name', '')
                    expense['timestamp'] = timestamp

                    # Insert the new stock entry into MongoDB
                    db.stock_expenses.insert_one(expense)
                    db.audit_logs.insert_one({
                        'user': login_data,
                        'Activity': 'Added new expense',
                        'Item': expense['expenseName'],
                        'timestamp': datetime.now()
                    })
                    flash('Expense(s) were added', 'success')
                except (ValueError, TypeError) as e:
                    # Log or handle the exception as needed
                    flash(f"Error processing expense {expense.get('expenseName', 'unknown')}: {e}", 'error')

            return jsonify({'redirect': url_for('stockManagement_route.expenses_page')})
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

###viewing stock history
@stockManagement.route('/view-expenses')
def view_expenses():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})
            if company.get('update_sales') in ('yes', None):
                subscription = db.managers.find_one({'name': company['company_name']}, {'account_type': 1, 'manager_email': 1, '_id': 0})
                account_type = subscription['account_type']
                # Remove any empty strings from the list
                account_type = [atype for atype in account_type if atype]

                if 'Enterprise Resource Planning' in account_type:
                    company_name = company['company_name']
                    now = datetime.now()
                    first_day_of_current_month = datetime(now.year, now.month, 1)
                    twelve_months_ago = first_day_of_current_month.replace(year=first_day_of_current_month.year - 1)
                    expense_info = list(db.stock_expenses.find({'company_name': company_name, 'expenseDate': {'$gte': twelve_months_ago}}))
                    expense_info.sort(key=lambda x: x.get('timestamp', x['expenseDate']), reverse=True)

                    dp = company.get('dp')
                    dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
                    return render_template('view expenses.html', expense_info = expense_info, dp=dp_str)
            else:
                flash('You do not have rights to view expenses', 'error')
                return redirect('/stock-details')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

###DOANLOAD EXPENSE DATA   
@stockManagement.route('/download-expense-data', methods=["POST"])
def download_expense_data():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            startdate_on_str = request.form.get("startdate")
            enddate_on_str = request.form.get("enddate")
            startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
            enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')

            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})

            expenses = list(db.stock_expenses.find(
                {'company_name': company['company_name'], 'expenseDate': {'$gte': startdate, '$lte': enddate}},
                {'_id': 0, 'company_name': 0}
            ))

            # Sort data by expenseDate in descending order
            sorted_expenses = sorted(expenses, key=lambda x: x["expenseDate"], reverse=True)

            # Create Excel file
            excel_buffer = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Expenses"

            # Write header row
            headers = ['Expense', 'Date', 'Amount']
            ws.append(headers)

            # Write data rows
            for expense in sorted_expenses:
                row = [
                    expense.get('expenseName', ''),
                    expense.get('expenseDate', '').strftime('%Y-%m-%d') if isinstance(expense.get('expenseDate'), datetime) else '',
                    expense.get('amount', 0)
                ]
                ws.append(row)

            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Create the response
            response = make_response(excel_buffer.getvalue())
            response.headers['Content-Disposition'] = f"attachment; filename={company['company_name']}_Expenses_{startdate_on_str}_{enddate_on_str}.xlsx"
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            # Clean up
            del wb
            del excel_buffer
            gc.collect()

            return response
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
####edit expense
@stockManagement.route('/edit-expense/<item_id>', methods=['GET', 'POST'])
def edit_expense(item_id):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            manager = db.registered_managers.find_one({'username':login_data},{'_id':0,'createdAt':0,'code':0,'address':0})
            if manager.get('update_sales') in ('yes', None):
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                selected_item = db.stock_expenses.find_one({'_id': ObjectId(item_id)})
                return render_template('edit-expense.html',item_id=item_id,dp=dp_str,selected_item=selected_item)
            else:
                flash('You do not have rights to edit', 'error')
                return redirect('/stock-details')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@stockManagement.route('/apply-expense-edits', methods=['POST'])
def apply_expense_edits():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            item_id = request.form.get("item_id")
            expense_name = request.form.get("expense_name")
            amount = request.form.get("amount")
            expensedate = request.form.get("expensedate")

            selected_item = db.stock_expenses.find_one({'_id': ObjectId(item_id)})

            fields_to_update = {}
            if selected_item:
                if expense_name:
                    fields_to_update['expenseName'] = expense_name
                if amount:
                    fields_to_update['amount'] = float(amount)
                if expensedate:
                    expensedate = datetime.strptime(expensedate, '%Y-%m-%d')
                    fields_to_update['expenseDate'] = expensedate
            else:
                flash('Please select an up-to-date expense', 'error')
            
            if not fields_to_update:
                flash('No edits were applied', 'error')
            else:
                db.stock_expenses.update_one({'_id': ObjectId(item_id)},
                                    {'$set': fields_to_update})
                db.audit_logs.insert_one({'user': login_data,'Activity': 'Edit expense','Item': item_id,'timestamp': datetime.now()})
                flash('Expense updates were applied', 'success')
            return redirect('/view-expenses')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

####delete expense
@stockManagement.route('/delete-expense/<item_id>', methods=['POST'])
def delete_expense(item_id):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            manager = db.registered_managers.find_one({'username':login_data},{'_id':0,'createdAt':0,'code':0,'address':0})
            if manager.get('update_sales') in ('yes', None):
                selected_item = db.stock_expenses.find_one({'_id': ObjectId(item_id)})
                if selected_item:
                    db.stock_expenses.delete_one({'_id': ObjectId(item_id)})
                    db.audit_logs.insert_one({'user': login_data,'Activity': 'Expense deletion','Item': item_id,'timestamp': datetime.now()})
                    flash('Expense was deleted', 'success')
                else:
                    flash('Expense does not exist', 'error')
            else:
                flash('You do not have rights to delete', 'error')
            return redirect('/view-expenses')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

####edit items
@stockManagement.route('/edit-item/<item_id>', methods=['GET', 'POST'])
def edit_item(item_id):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            manager = db.registered_managers.find_one({'username':login_data},{'_id':0,'createdAt':0,'code':0,'address':0})
            if manager.get('update_stock') in ('yes', None):
                selected_item = db.inventories.find_one({'_id': ObjectId(item_id)})
                if selected_item:
                    company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                    if 'dp' in company:
                        dp_str = company['dp']
                    else:
                        dp_str = None
                    return render_template('edit-stock.html',item_id=item_id,dp=dp_str,selected_item=selected_item)
                else:
                    flash('Please select an up-to-date item', 'error')
                    return redirect('/stock-details')
            else:
                flash('You do not have rights to edit', 'error')
                return redirect('/stock-details')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@stockManagement.route('/apply-item-edits', methods=['POST'])
def apply_item_edits():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
            
            item_id = request.form.get("item_id")
            item_name = request.form.get("item_name")
            quantity = request.form.get("quantity")
            unit_price = request.form.get("unit_price")
            stockdate = request.form.get("stockdate")
            unit_of_measurement = request.form.get("unit_of_measurement")

            selected_item = db.inventories.find_one({'_id': ObjectId(item_id)})

            applied = 0
            if selected_item:
                if item_name:
                    db.inventories.update_one({'company_name': company['company_name'],'itemName': selected_item['itemName']}, {'$set': {'itemName': item_name}})
                    db.old_inventories.update_many({'company_name': company['company_name'],'itemName': selected_item['itemName']}, {'$set': {'itemName': item_name}})
                    db.stock_sales.update_many({'company_name': company['company_name'],'itemName': selected_item['itemName']}, {'$set': {'itemName': item_name}})
                    applied = 1
                if unit_of_measurement:
                    db.inventories.update_one({'company_name': company['company_name'],'itemName': selected_item['itemName']}, {'$set': {'unitOfMeasurement': unit_of_measurement}})
                    applied = 1
                if quantity:
                    quantity = float(quantity)
                    if 'available_quantity' in selected_item:
                        new_qty = selected_item['available_quantity'] - selected_item['quantity']
                        if new_qty < 0:
                            new_qty = 0
                            flash('Item sales were already updated', 'error')
                        available_quantity = new_qty + quantity                        
                    else:
                        available_quantity = quantity
                    
                    totalPrice = selected_item['unitPrice'] * quantity
                    if 'cumulativeOldPrices' in selected_item:
                        cumulativeOldPrices = selected_item['cumulativeOldPrices'] - selected_item['totalPrice'] + totalPrice
                    else:
                        cumulativeOldPrices = totalPrice
                    db.inventories.update_one({'_id': ObjectId(item_id)}, {'$set': {'quantity': quantity, 'available_quantity': available_quantity, 'totalPrice': totalPrice, 'cumulativeOldPrices': cumulativeOldPrices}})
                    applied = 1
                if unit_price:
                    unit_price = float(unit_price)
                    if quantity:
                        new_total_price = quantity * unit_price
                    else:
                        new_total_price = selected_item['quantity'] * unit_price
                    
                    if 'cumulativeOldPrices' in selected_item:
                        cumulativeOldPrices = selected_item['cumulativeOldPrices'] - selected_item['totalPrice'] + new_total_price
                    else:
                        cumulativeOldPrices = new_total_price
                    db.inventories.update_one({'_id': ObjectId(item_id)}, {'$set': {'unitPrice': unit_price, 'totalPrice': new_total_price, 'cumulativeOldPrices': cumulativeOldPrices}})
                    applied = 1
                if stockdate:
                    stockDate = datetime.strptime(stockdate, '%Y-%m-%d')
                    db.inventories.update_one({'_id': ObjectId(item_id)}, {'$set': {'stockDate': stockDate}})
                    applied = 1

                if applied == 1:
                    flash('Item updates were applied', 'success')
                    db.audit_logs.insert_one({'user': login_data,'Activity': 'Stock edit','Item': item_id,'timestamp': datetime.now()})
                else:
                    flash('No edits were made', 'error')
                return redirect('/stock-details')
            else:
                flash('Please select an up-to-date item', 'error')
                return redirect('/stock-details')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
        
####delete items
@stockManagement.route('/delete-item/<item_id>', methods=['POST'])
def delete_item(item_id):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            manager = db.registered_managers.find_one({'username':login_data},{'_id':0,'createdAt':0,'code':0,'address':0})
            if manager.get('update_stock') in ('yes', None):
                selected_item = db.inventories.find_one({'_id': ObjectId(item_id)})
                if selected_item:
                    db.inventories.delete_one({'_id': ObjectId(item_id)})
                    db.audit_logs.insert_one({'user': login_data,'Activity': 'Stock deletion','Item': item_id,'timestamp': datetime.now()})
                    flash('Item was deleted', 'success')
                else:
                    flash('Please select an up-to-date item', 'error')
            else:
                flash('You do not have rights to delete', 'error')
            return redirect('/stock-details')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
####delete sale
@stockManagement.route('/delete-sale/<item_id>', methods=['POST'])
def delete_sale(item_id):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            manager = db.registered_managers.find_one({'username':login_data},{'_id':0,'createdAt':0,'code':0,'address':0})
            if manager.get('update_sales') in ('yes', None):
                sale_to_delete = db.stock_sales.find_one({'_id': ObjectId(item_id)})
                if sale_to_delete:
                    if 'stock_id' in sale_to_delete:
                        stock_to_undo = db.inventories.find_one({'_id': sale_to_delete['stock_id']})
                        if stock_to_undo:
                            available_quantity = stock_to_undo['available_quantity'] + sale_to_delete['quantity']
                            db.inventories.update_one({'_id': sale_to_delete['stock_id']}, {'$set': {'available_quantity': available_quantity}})
                            db.stock_sales.delete_one({'_id': ObjectId(item_id)})
                            db.audit_logs.insert_one({'user': login_data,'Activity': 'Sale deletion','Item': item_id,'timestamp': datetime.now()})
                            flash('Sale was deleted', 'success')
                        else:
                            flash('Unable to delete: No stock available', 'error')
                    else:
                        stock_to_undo = db.inventories.find_one({'company_name': manager['company_name'],'itemName': sale_to_delete['itemName']})
                        if stock_to_undo:
                            available_quantity = stock_to_undo['available_quantity'] + sale_to_delete['quantity']
                            db.inventories.update_one({'company_name': manager['company_name'],'itemName': sale_to_delete['itemName']}, {'$set': {'available_quantity': available_quantity}})
                            db.stock_sales.delete_one({'_id': ObjectId(item_id)})
                            db.audit_logs.insert_one({'user': login_data,'Activity': 'Sale deletion','Item': item_id,'timestamp': datetime.now()})
                            flash('Sale was deleted', 'success')
                        else:
                            flash('Unable to delete: No stock available', 'error')
                else:
                    flash('Sale does not exist', 'error')
            else:
                flash('You do not have rights to delete', 'error')
            return redirect('/sales-details')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

def remove_file_later(filepath, delay=10):
    """ Schedule file deletion after a delay """
    def delayed_removal():
        time.sleep(delay)
        if os.path.exists(filepath):
            os.remove(filepath)
    threading.Thread(target=delayed_removal).start()

@stockManagement.route('/store-bar-code', methods=['POST'])
def store_bar_code():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Enterprise Resource Planning':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
            
            product_name = request.form.get('typed_input')
            product_id = db.inventories.find_one({'company_name': company['company_name'],'itemName': product_name})
            product_id_string = product_id['product_id']

            # Generate the barcode using the default options
            CODE128 = barcode.get_barcode_class('code128')
            writer = ImageWriter()

            # Set writer options to disable default text and reduce bar height
            options = {
                'write_text': False,   # Disable the default text below the barcode
                'module_height': 10.0  # Set the height of the bars to a smaller value
            }

            # Generate the barcode image with the custom options
            barcode_image = CODE128(product_id_string, writer=writer)
            barcode_img = barcode_image.render(writer_options=options)

            # Convert to Pillow Image to add text
            draw = ImageDraw.Draw(barcode_img)

            # Use a default font from Pillow
            font = ImageFont.load_default()

            # Add custom text below the barcode
            img_width, img_height = barcode_img.size
            text_width, text_height = draw.textbbox((0, 0), product_id_string, font=font)[2:]  # Corrected line using textbbox
            text_position = ((img_width - text_width) // 2, img_height - text_height)

            draw.text(text_position, product_id_string, font=font, fill="black")

            # Create filename and filepath
            filename = f'{product_name}.png'
            filepath = os.path.join('.', filename)

            # Save the barcode image to a file
            barcode_img.save(filepath, 'PNG')

            # Convert the image to bytes for response
            img_io = io.BytesIO()
            barcode_img.save(img_io, 'PNG')
            img_io.seek(0)

            # Flash success message
            flash(f'Barcode for {product_name} was generated and downloaded to your device', 'success')
            # Schedule file removal after a delay
            remove_file_later(filepath, delay=10)
            return jsonify({
                'download_url': url_for('stockManagement_route.download_barcode', filename=filename),
                'redirect_url': url_for('stockManagement_route.generate_bar_codes')
            })
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@stockManagement.route('/download-barcode/<filename>')
def download_barcode(filename):
    return send_from_directory(directory='.', path=filename, as_attachment=True, download_name=filename)

@stockManagement.route('/api/<api_key>/<data>', methods=['GET'])
def get_data(api_key, data):
    result = []
    if api_key and data:
        db, fs = get_db_and_fs()
        try:
            api = db.managers.find_one({'_id': ObjectId(api_key)})
        except Exception as e:
            return jsonify({"error": "Invalid API Key"}), 400

        if api:
            if data == 'stock':
                current_stock = db.inventories.find(
                    {'company_name': api['name']},
                    {'_id': 0, 'company_name': 0}
                )
                old_stock = db.old_inventories.find(
                    {'company_name': api['name']},
                    {'_id': 0, 'company_name': 0}
                )
                combined_stock = list(current_stock) + list(old_stock)

                sorted_combined_stock = sorted(combined_stock, key=lambda x: x.get("stockDate", ""), reverse=True)
                for item in sorted_combined_stock:
                    result.append({
                        'Item Name': item['itemName'],
                        'Stock Date': item['stockDate'],
                        'Stocked Quantity': item['quantity'],
                        'Available Stock': item['available_quantity'],
                        'Measurement': item['unitOfMeasurement'],
                        'Buying Price': item['unitPrice'],
                        'Total Buying Price': item['totalPrice']
                    })
            elif data == 'sales':
                sales_info = list(db.stock_sales.find({
                'company_name': api['name']
                }, {
                    '_id': 0,
                    'company_name': 0,
                    'stockDate': 0
                }))

                sorted_sales_info = sorted(sales_info, key=lambda x: x["saleDate"], reverse=True)

                for item in sorted_sales_info:
                    result.append({
                        'Item Name': item['itemName'],
                        'Sale Date': item['saleDate'],
                        'Sold Quantity': item['quantity'],
                        'Selling Price': item['unitPrice'],
                        'Total Selling Price': item['revenue'],
                    })
            elif data == 'profits':
                pipeline = [
                    {
                        '$match': {
                            'company_name': api['name']
                        }
                    },
                    {
                        '$group': {
                            '_id': {'itemName': '$itemName', 'stockDate': '$stockDate'},
                            'totalRevenue': {'$sum': '$revenue'},
                            'quantitySold': {'$sum': '$quantity'}
                        }
                    },
                    {
                        '$lookup': {
                            'from': 'inventories',
                            'let': {'itemName': '$_id.itemName', 'stockDate': '$_id.stockDate'},
                            'pipeline': [
                                {
                                    '$match': {
                                        '$expr': {
                                            '$and': [
                                                {'$eq': ['$itemName', '$$itemName']},
                                                {'$eq': ['$company_name', api['name']]},
                                                { '$eq': ['$stockDate', '$$stockDate'] }
                                            ]
                                        }
                                    }
                                },
                                {
                                    '$project': {
                                        '_id': 0,
                                        'quantity': 1,
                                        'unitPrice': 1,
                                        'stockDate': 1
                                    }
                                }
                            ],
                            'as': 'inventoryDetails'
                        }
                    },
                    {
                        '$lookup': {
                            'from': 'old_inventories',
                            'let': {'itemName': '$_id.itemName', 'stockDate': '$_id.stockDate'},
                            'pipeline': [
                                {
                                    '$match': {
                                        '$expr': {
                                            '$and': [
                                                {'$eq': ['$itemName', '$$itemName']},
                                                {'$eq': ['$company_name', api['name']]},
                                                { '$eq': ['$stockDate', '$$stockDate'] }
                                            ]
                                        }
                                    }
                                },
                                {
                                    '$project': {
                                        '_id': 0,
                                        'quantity': 1,
                                        'unitPrice': 1,
                                        'stockDate': 1
                                    }
                                }
                            ],
                            'as': 'oldInventoryDetails'
                        }
                    },
                    {
                        '$project': {
                            'inventoryDetails': {
                                '$cond': {
                                    'if': {'$gt': [{'$size': '$inventoryDetails'}, 0]},
                                    'then': '$inventoryDetails',
                                    'else': '$oldInventoryDetails'
                                }
                            },
                            'totalRevenue': 1,
                            'quantitySold': 1,
                            '_id': '$_id'
                        }
                    },
                    {
                        '$unwind': '$inventoryDetails'
                    },
                    {
                        '$group': {
                            '_id': '$_id.itemName',
                            'stockDate': {'$first': '$_id.stockDate'},
                            'totalRevenue': {'$first': '$totalRevenue'},
                            'quantitySold': {'$first': '$quantitySold'},
                            'unitPrice': {'$avg': '$inventoryDetails.unitPrice'}
                        }
                    },
                    {
                        '$addFields': {
                            'unitProfit': {
                                '$round': [
                                    {
                                        '$subtract': [
                                            {'$divide': ['$totalRevenue', '$quantitySold']},
                                            '$unitPrice'
                                        ]
                                    },
                                    0
                                ]
                            },
                            'totalProfit': {
                                '$subtract': [
                                    '$totalRevenue',
                                    {'$multiply': ['$unitPrice', '$quantitySold']}
                                ]
                            }
                        }
                    },
                    {
                        '$match': {
                            'quantitySold': {'$ne': 0}
                        }
                    }
                ]

                revenue_info = list(db.stock_sales.aggregate(pipeline))
                revenue_info.sort(key=lambda x: x['_id'])

                if revenue_info:
                    for revenue in revenue_info:
                        item_name = revenue['_id']
                        stock_date = revenue['stockDate'].strftime('%Y-%m-%d')
                        buying_price = revenue['unitPrice']
                        quantity_sold = revenue['quantitySold']
                        total_revenue = revenue['totalRevenue']                   

                        unitProfit = round(revenue['unitProfit'],0)
                        totalProfit = revenue['totalProfit']

                        result.append({
                            'Item Name': item_name,
                            'Stock Date': stock_date,
                            'Buying Price': buying_price,
                            'Sold Quantity': quantity_sold,
                            'Total Selling Price': total_revenue,
                            'Average Unit Profit': unitProfit,
                            'Total Profit': totalProfit
                        })
            elif data == 'expenses':
                expenses = list(db.stock_expenses.find(
                    {'company_name': api['name']},
                    {'_id': 0, 'company_name': 0}
                ))

                sorted_expenses = sorted(expenses, key=lambda x: x["expenseDate"], reverse=True)

                for expense in sorted_expenses:
                    result.append({
                        'Expense': expense.get('expenseName', ''),
                        'Date': expense.get('expenseDate', '').strftime('%Y-%m-%d') if isinstance(expense.get('expenseDate'), datetime) else '',
                        'Amount': expense.get('amount', 0)
                    })
            else:
                return jsonify({"error": "Invalid data request"}), 400
        else:
            return jsonify({"error": "API Key not found"}), 404
    else:
        return jsonify({"error": "Missing API Key or Data"}), 400
    
    return jsonify(result)