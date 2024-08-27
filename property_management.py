from flask import Blueprint, render_template, url_for, send_from_directory, request, flash, redirect, session, make_response, jsonify, current_app
from utils import get_mongo_client, get_db_and_fs, send_async_email
from flask_mail import Message
import threading
from docx import Document
from pymongo import MongoClient, ASCENDING, DESCENDING
import secrets
import bcrypt
from datetime import datetime, timedelta, timezone
import calendar
import pytz
import pandas as pd 
from io import BytesIO
import json
from bson.objectid import ObjectId
import cv2
import numpy as np
import io
import base64
import random
import os
from werkzeug.utils import secure_filename
from gridfs import GridFS
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import ZipFile
import tempfile
import string
import qrcode
import time
from docx2pdf import convert
import PyPDF2
import gc
from collections import defaultdict
import barcode
from PIL import Image, ImageDraw, ImageFont
from barcode.writer import ImageWriter

propertyManagement = Blueprint('propertyManagement_route', __name__)

def parse_iso_format(iso_str):
    if iso_str.endswith('Z'):
        iso_str = iso_str[:-1]  # Remove 'Z'
        dt = datetime.fromisoformat(iso_str)
        return dt.replace(tzinfo=pytz.UTC)
    else:
        return datetime.fromisoformat(iso_str)

def generate_file_password(length=12):
    characters = string.ascii_letters + string.digits
    return ''.join(random.choice(characters) for _ in range(length))

def get_managers_data(registered_managers):
    managers = []
    for manager in registered_managers:
        managers.append((manager['name'], manager['email'], manager['phone_number'], manager['company_name']))
    return managers

#############DASHBOARD PAGE#######################
@propertyManagement.route('/load-dashboard-page', methods=["GET", "POST"])
def load_dashboard_page():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    # get the current month number
    current_month = datetime.now().month
    # convert the month number to its name
    month_name = calendar.month_name[current_month]
    # capitalize the first letter
    month_name = month_name.capitalize()

    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
            if session.get('is_manager') != "is_manager":
                flash("You do not have rights to view the dashboard","error")
                dp = company.get('dp')
                dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
                return render_template('dashboard.html', chart_property_performance_trended_data=[],chart_property_performance_data=[],chart_property_type_data=[],dp=dp_str)
            else:
                send_emails = db.send_emails.find_one({'emails': "yes"},{'emails': 1})

                if send_emails is None:
                    session['send_emails_message'] = "Our email service is currently unavailable. We apologize for any inconvenience. Our team is working hard to fix the issue and we expect the service to be back soon."
                        
                startdate_on_str = request.form.get("startdate")
                enddate_on_str = request.form.get("enddate")

                month_mapping = {
                    'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
                    'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12,
                    'Quarter 1': 3, 'Quarter 2': 6, 'Quarter 3': 9, 'Quarter 4': 12,
                    '2024': 12, '2025': 12, '2026': 12
                }

                subscription = db.managers.find_one({'name': company['company_name']}, {'account_type': 1, 'manager_email': 1, '_id': 0})

                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None

                if subscription['manager_email'] == company['email']:
                    user_query  = {'company_name': company['company_name']}
                else:
                    user_query  = {'username': login_data, 'company_name': company['company_name']}
                    
                projection = {'payment_receipt': 0, 'username': 0, 'company_name': 0, 'tenantEmail': 0, 'tenantPhone': 0, 'payment_mode': 0,
                            'gender': 0, 'household_size': 0, 'payment_type': 0, 'payment_completion': 0, 'currency': 0, 'payment_status': 0,
                            '_id': 0}
                property_data_list = list(db.property_managed.find(user_query))
                if len(property_data_list) == 0:
                    flash('No property data found', 'error')
                    return render_template('dashboard.html', chart_property_performance_trended_data=[],chart_property_performance_data=[],chart_property_type_data=[],dp=dp_str)
                else:
                    if startdate_on_str and enddate_on_str:
                        startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
                        enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')
                        latest_year = enddate.year

                        date_query = {'date_last_paid': {'$gte': startdate, '$lte': enddate}, 'status': {'$ne': 'deleted'}}
                        date_query.update(user_query)

                        current_tenant_data = list(db.tenants.find(date_query, projection))

                        old_tenant_data = list(db.old_tenant_data.find(date_query, projection))

                        tenant_data_stats = list(db.tenants.find({}, projection))

                        month_name = f"{startdate_on_str} to {enddate_on_str}"
                    else:
                        latest_document = db.tenants.find_one(sort=[('date_last_paid', -1)], projection={'date_last_paid': 1, '_id': 0})
                        if latest_document is None:
                            latest_document_old = db.old_tenant_data.find_one(sort=[('date_last_paid', -1)], projection={'date_last_paid': 1, '_id': 0})
                            if latest_document_old is None:
                                flash('No tenant data found', 'error')
                                return render_template('dashboard.html', chart_property_performance_trended_data=[],chart_property_performance_data=[],chart_property_type_data=[],dp=dp_str)
                            else:
                                latest_year = latest_document_old['date_last_paid'].year
                        else:    
                            latest_year = latest_document['date_last_paid'].year

                        startdate = datetime(latest_year, 1, 1)
                        enddate = datetime(latest_year, 12, 31, 23, 59, 59)

                        date_query = {'date_last_paid': {'$gte': startdate, '$lte': enddate}, 'status': {'$ne': 'deleted'}}
                        date_query.update(user_query)

                        current_tenant_data = list(db.tenants.find(date_query, projection))

                        old_tenant_data = list(db.old_tenant_data.find(date_query, projection))

                        tenant_data_stats = list(db.tenants.find({}, projection))

                        month_name = f"{startdate.strftime('%Y-%m-%d')} to {enddate.strftime('%Y-%m-%d')}"
                    
                    overdue_tenants = []
                    count_current_tenants = 0
                    if len(tenant_data_stats) != 0:
                        for count_tenant in tenant_data_stats:
                            if count_tenant['status'] != 'deleted':
                                count_current_tenants += 1

                                last_payment_month = month_mapping.get(count_tenant['months_paid'], 0)
                                last_payment_date = datetime(year=count_tenant['year'], month=last_payment_month, day=1)
                                next_payment_date = last_payment_date + timedelta(days=30)
                                remaining_days = (next_payment_date - datetime.now()).days
                                if remaining_days < 0:
                                    overdue = True
                                    remaining_days = abs(remaining_days)
                                    if remaining_days < 7:
                                        time_unit = 'day(s)'
                                    elif remaining_days < 30:
                                        remaining_days = round(remaining_days / 7)
                                        time_unit = 'week(s)'
                                    elif remaining_days < 365:
                                        remaining_days = round(remaining_days / 30)
                                        time_unit = 'month(s)'
                                    else:
                                        remaining_days = round(remaining_days / 365)
                                        time_unit = 'year(s)'

                                    overdue_status = 'overdue' if overdue else 'due in'
                                    overdue_tenants.append((count_tenant['tenantName'], count_tenant['propertyName'],count_tenant['selected_section'], remaining_days, time_unit, overdue_status))

                        overdue_tenants = sorted(overdue_tenants, key=lambda x: x[3], reverse=True)

                    doc_query = {'status': {'$ne': 'deleted'}}
                    doc_query.update(user_query)
                    tenant_data_cursor = db.tenants.find(doc_query, projection)
                    
                    property_data_dict = {doc['propertyName']: doc['sections'] for doc in property_data_list}
                    property_occupancy = {doc['propertyName']: {'total': len(doc['sections']), 'occupied': 0} for doc in property_data_list}
                    for tenant_exists in tenant_data_cursor:
                        tenant_property_name = tenant_exists.get('propertyName', '').strip()
                        selected_section = tenant_exists.get('selected_section', '').strip()
                        # Check if the property exists in updated_property_data and the section is in the property's sections
                        if tenant_property_name in property_data_dict and selected_section in property_data_dict[tenant_property_name]:
                            # Remove the section
                            property_data_dict[tenant_property_name].remove(selected_section)
                            # Increase the count of occupied sections
                            property_occupancy[tenant_property_name]['occupied'] += 1
                            # If there are no more sections for this property, remove the property
                            if not property_data_dict[tenant_property_name]:
                                del property_data_dict[tenant_property_name]

                    if current_tenant_data or old_tenant_data:  # Check if data is available
                        combined_data = current_tenant_data + old_tenant_data

                        # Create a mapping of property values
                        property_value_dict = {item['propertyName']: item['property_value'] for item in property_data_list}

                        # Initialize dictionaries to hold data
                        property_performance = defaultdict(lambda: {
                            'available_amount': 0,
                            'months_paid': set(),
                            'property_value': 0,
                            'demanded_amount': 0
                        })
                        
                        # Process combined_data
                        for item in combined_data:
                            prop_name = item['propertyName']
                            if prop_name in property_value_dict:
                                property_value = property_value_dict[prop_name]
                                property_performance[prop_name]['available_amount'] += item.get('available_amount', 0)
                                property_performance[prop_name]['months_paid'].add(item.get('months_paid', ''))
                                property_performance[prop_name]['property_value'] = property_value
                        
                        # Compute months_paid_count and total_property_value
                        for prop_name, data in property_performance.items():
                            data['months_paid_count'] = len(data['months_paid'])
                            data['total_property_value'] = data['months_paid_count'] * data['property_value']
                            data['demanded_amount'] = data['total_property_value'] - data['available_amount']
                        
                        # Initialize for property_performance_line
                        property_performance_line = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
                        for item in combined_data:
                            year = item.get('year')
                            month = item.get('months_paid')
                            prop_name = item['propertyName']
                            property_performance_line[year][month][prop_name] += item.get('available_amount', 0)
                        
                        # Prepare data for property_performance_line for the chart
                        month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
                        labels = month_order
                        datasets = {prop_name: [property_performance_line[year].get(month, {}).get(prop_name, 0) for month in month_order] for prop_name in property_performance.keys()}

                        chart_property_performance_trended_data = {
                            'labels': labels,
                            'datasets': [{'label': prop_name, 'data': datasets[prop_name]} for prop_name in property_performance.keys()]
                        }

                        chart_property_performance_data = {
                            'labels': list(property_performance.keys()),
                            'available_amount': [data['available_amount'] for data in property_performance.values()],
                            'demanded_amount': [data['demanded_amount'] for data in property_performance.values()]
                        }

                        # Count properties and available amount
                        count_property = len(property_data_list)
                        available_amount = sum(item.get('available_amount', 0) for item in combined_data)

                        # Group property data
                        property_type_counts = defaultdict(int)
                        for item in property_data_list:
                            property_type_counts[item['type']] += 1
                        
                        chart_property_type_data = {
                            'labels': list(property_type_counts.keys()),
                            'values': list(property_type_counts.values())
                        }

                        return render_template('dashboard.html', 
                                            chart_property_performance_trended_data=chart_property_performance_trended_data,
                                            chart_property_performance_data=chart_property_performance_data,
                                            chart_property_type_data=chart_property_type_data, 
                                            count_property=count_property,
                                            available_amount=available_amount,
                                            overdue_tenants=overdue_tenants,
                                            property_occupancy=property_occupancy,
                                            count_current_tenants=count_current_tenants,
                                            month_name=month_name, dp=dp_str)
                    else:
                        flash('No tenant data found', 'error')
                        return render_template('dashboard.html', chart_property_performance_trended_data=[],chart_property_performance_data=[],chart_property_type_data=[],dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

#Manager notifications
@propertyManagement.route('/manager notifications')
def manager_notifications():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        notifications = []
        timestamps = []

        tenants = list(db.tenant_user_accounts.find({'account_manager': login_data}, {'_id': 1}))
        if len(tenants) != 0:
            for tenant in tenants:
                # Retrieve new complaints
                new_complaints = list(db.tenant_complaints.find(
                    {'tenantID': tenant['_id']}
                ))

                if new_complaints:
                    for new_complaint in new_complaints:
                        complained_on = new_complaint['complained_on']
                        if isinstance(complained_on, datetime):
                            formatted_complaint_date = complained_on.strftime('%Y-%m-%d %H:%M')
                        else:
                            complained_on = datetime.fromisoformat(complained_on)
                            formatted_complaint_date = complained_on.strftime('%Y-%m-%d %H:%M')

                        notification = f"Complaint from {new_complaint['tenant_name']} on {formatted_complaint_date}"
                        notifications.append(notification)
                        timestamps.append(complained_on)

                    # Retrieve and sort replies, limiting to 20
                    replies = list(db.tenant_complaints_replies.find(
                        {'complaintID': {'$in': [complaint['_id'] for complaint in new_complaints]}}
                    ).sort('reply_date', DESCENDING).limit(20))

                    if replies:
                        for reply in replies:
                            reply_date = reply['reply_date']
                            if isinstance(reply_date, datetime):
                                formatted_reply_date = reply_date.strftime('%Y-%m-%d %H:%M')
                            else:
                                reply_date = datetime.fromisoformat(reply_date)
                                formatted_reply_date = reply_date.strftime('%Y-%m-%d %H:%M')
                            
                            if reply['who'] != 'Manager':
                                notification = f"Reply from {reply['who']} on {formatted_reply_date}"
                                notifications.append(notification)
                                timestamps.append(reply_date)

        # Combine notifications with their timestamps and sort
        combined = list(zip(notifications, timestamps))
        combined.sort(key=lambda x: x[1], reverse=True)  # Sort by timestamp, latest first

        # Separate sorted notifications and timestamps
        notifications = [notif for notif, _ in combined]

        company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
        dp = company.get('dp')
        dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
        
        return render_template('manager_notifications.html', notifications=notifications, dp=dp_str)

####ASSIGN PROPERTIES TO MANAGERS
@propertyManagement.route('/assign-properties')
def assign_properties():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            if 'dp' in company:
                dp_str = company['dp']
            else:
                dp_str = None
            
            # Get registered managers data
            registered_managers = list(db.registered_managers.find({'company_name': company['company_name'], 'username': {'$ne': login_data}}))
            if not registered_managers:
                flash("We did not find other registered users", 'error')
                return redirect('/load-dashboard-page')

            # Prepare managers data
            managers = get_managers_data(registered_managers)

            return render_template('assign properties.html',managers=managers,dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@propertyManagement.route('/assign-properties-page/<name>/<email>/<company_name>')
def assign_properties_page(name,email,company_name):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            if 'dp' in company:
                dp_str = company['dp']
            else:
                dp_str = None
            properties = db.property_managed.find({'company_name': company['company_name']}, {"propertyName": 1})
            property_names = [property['propertyName'] for property in properties]
            
            return render_template('assign properties page.html', property_names=property_names,name=name,email=email,company_name=company_name,dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@propertyManagement.route('/assign-properties-initiated', methods=["POST"])
def assign_properties_initiated():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            name = request.form.get('name')
            email = request.form.get('email')
            company_name = request.form.get('company_name')
            propertyName = request.form.get("propertyName")

            db.registered_managers.update_one(
                {'email': email, 'company_name': company_name}, 
                {'$addToSet': {'properties': propertyName}},
                upsert=True
            )
            db.audit_logs.insert_one({'user': login_data, 'Activity': 'Assign property', 'email':email, 'timestamp': datetime.now()})
            flash(f"{propertyName} was assigned to {name}", 'success')
            return redirect('/assign-properties')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

####UNASSIGN PROPERTIES FROM MANAGERS
@propertyManagement.route('/unassign-properties')
def unassign_properties():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            if 'dp' in company:
                dp_str = company['dp']
            else:
                dp_str = None
            
            # Get registered managers data
            registered_managers = list(db.registered_managers.find({'company_name': company['company_name'], 'username': {'$ne': login_data}}))
            if not registered_managers:
                flash("We did not find other registered users", 'error')
                return redirect('/load-dashboard-page')

            # Prepare managers data
            managers = get_managers_data(registered_managers)

            return render_template('unassign properties.html',managers=managers,dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@propertyManagement.route('/unassign-properties-page/<name>/<email>/<company_name>')
def unassign_properties_page(name,email,company_name):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            if 'dp' in company:
                dp_str = company['dp']
            else:
                dp_str = None
            property_assigned = db.registered_managers.find({'email': email, 'company_name': company_name})
            property_assigned_dict = {property for doc in property_assigned if 'properties' in doc for property in doc['properties']}
            
            return render_template('unassign properties page.html', property_names=property_assigned_dict,name=name,email=email,company_name=company_name,dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@propertyManagement.route('/unassign-properties-initiated', methods=["POST"])
def unassign_properties_initiated():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            name = request.form.get('name')
            email = request.form.get('email')
            company_name = request.form.get('company_name')
            propertyName = request.form.get("propertyName")

            db.registered_managers.update_one(
                {'email': email, 'company_name': company_name}, 
                {'$pull': {'properties': propertyName}}
            )
            db.audit_logs.insert_one({'user': login_data, 'Activity': 'Unassign property', 'email':email, 'timestamp': datetime.now()})
            flash(f"{propertyName} was unassigned from {name}", 'success')
            return redirect('/unassign-properties')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@propertyManagement.route('/add properties')
def add_properties():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            if session.get('add_properties') == "no":
                flash("You do not have rights to add properties","error")
                return redirect("/load-dashboard-page")
            else:
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                return render_template('add property page.html', dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@propertyManagement.route('/add tenants')
def add_tenants():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            if session.get('add_tenants') == "no":
                flash("You do not have rights to add tenants","error")
                return redirect("/load-dashboard-page")
            else:
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                is_manager = db.managers.find_one({'manager_email': company['email'], 'name':company['company_name']})        

                if is_manager is None:
                    user_query  = {'username': login_data, 'company_name': company['company_name']}
                else:
                    user_query  = {'company_name': company['company_name']}

                property_data_list = list(db.property_managed.find(user_query,{'propertyName':1,'sections':1,'_id':0}))
                tenant_data_cursor = db.tenants.find(user_query,{'propertyName':1,'selected_section':1,'_id':0})
                        
                property_data_dict = {doc['propertyName']: doc['sections'] for doc in property_data_list}
                for tenant_exists in tenant_data_cursor:
                    tenant_property_name = tenant_exists.get('propertyName', '').strip()
                    selected_section = tenant_exists.get('selected_section', '').strip()
                    # Check if the property exists in updated_property_data and the section is in the property's sections
                    if tenant_property_name in property_data_dict and selected_section in property_data_dict[tenant_property_name]:
                        # Remove the section
                        property_data_dict[tenant_property_name].remove(selected_section)
                        # If there are no more sections for this property, remove the property
                        if not property_data_dict[tenant_property_name]:
                            del property_data_dict[tenant_property_name]
                return render_template('add tenants page.html', dp=dp_str, property_data=property_data_dict)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@propertyManagement.route('/export tenant data')
def export_tenant_data():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            if 'dp' in company:
                dp_str = company['dp']
            else:
                dp_str = None
            return render_template('export tenant data.html', dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

######add expenses
@propertyManagement.route('/property-expenses-page')
def property_expenses():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            
            if company.get('add_properties') in ('yes', None):
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                
                property_data_list = list(db.property_managed.find({'company_name': company['company_name']},{'propertyName':1,'_id':0}))

                return render_template('property expenses.html', dp=dp_str,property_data=property_data_list)
            else:
                flash('You do not have rights to add expenses', 'error')
                return redirect("/load-dashboard-page")
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@propertyManagement.route('/add-new-property-expense', methods=['POST'])
def add_new_property_expense():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                    'password': 0, 'auth': 0, 'dark_mode': 0})
                
            all_items = request.json.get('items', [])  # Access the JSON data sent from the client
            timestamp = datetime.now()

            for expense in all_items:
                expense['expenseName'] = expense.get('expenseName', '').strip()
                
                try:
                    # Convert 'quantity' and 'unitPrice' to floats
                    expense['amount'] = float(expense.get('amount', 0))
                    expense['expenseDate'] = datetime.strptime(expense.get('expenseDate', ''), '%Y-%m-%d')

                    expense['company_name'] = company.get('company_name', '')
                    expense['timestamp'] = timestamp

                    # Insert the new stock entry into MongoDB
                    db.property_expenses.insert_one(expense)
                    db.audit_logs.insert_one({
                        'user': login_data,
                        'Activity': 'Added new expense',
                        'Item': expense['expenseName'],
                        'timestamp': datetime.now()
                    })
                    flash('Expense(s) were added', 'success')
                except (ValueError, TypeError) as e:
                    # Log or handle the exception as needed
                    flash(f"Error processing expense {expense.get('expenseName', 'unknown')}: {e}", 'error')

            return jsonify({'redirect': url_for('propertyManagement_route.property_expenses')})
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

###viewing stock history
@propertyManagement.route('/view-property-expenses')
def view_property_expenses():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})
            if company.get('add_properties') in ('yes', None):
                company_name = company['company_name']
                twelve_months_ago = datetime.now() - timedelta(days=365)
                expense_info = list(db.property_expenses.find({'company_name': company_name, 'expenseDate': {'$gte': twelve_months_ago}}))
                expense_info.sort(key=lambda x: x.get('timestamp', x['expenseDate']), reverse=True)

                dp = company.get('dp')
                dp_str = base64.b64encode(base64.b64decode(dp)).decode() if dp else None
                return render_template('view property expenses.html', expense_info = expense_info, dp=dp_str)
            else:
                flash('You do not have rights to view expenses', 'error')
                return redirect("/load-dashboard-page")
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

####edit expense
@propertyManagement.route('/edit-property-expense/<item_id>', methods=['GET', 'POST'])
def edit_property_expense(item_id):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            manager = db.registered_managers.find_one({'username':login_data},{'_id':0,'createdAt':0,'code':0,'address':0})
            if manager.get('add_properties') in ('yes', None):
                if 'dp' in manager:
                    dp_str = manager['dp']
                else:
                    dp_str = None
                return render_template('edit property expenses.html',item_id=item_id,dp=dp_str)
            else:
                flash('You do not have rights to edit', 'error')
                return redirect("/load-dashboard-page")
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
    
@propertyManagement.route('/apply-property-expense-edits', methods=['POST'])
def apply_property_expense_edits():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            item_id = request.form.get("item_id")
            expense_name = request.form.get("expense_name")
            propertyName = request.form.get("propertyName")
            amount = request.form.get("amount")
            expensedate = request.form.get("expensedate")

            selected_item = db.property_expenses.find_one({'_id': ObjectId(item_id)})

            fields_to_update = {}
            if selected_item:
                if expense_name:
                    fields_to_update['expenseName'] = expense_name
                if propertyName:
                    fields_to_update['propertyName'] = propertyName
                if amount:
                    fields_to_update['amount'] = float(amount)
                if expensedate:
                    expensedate = datetime.strptime(expensedate, '%Y-%m-%d')
                    fields_to_update['expenseDate'] = expensedate
            else:
                flash('Please select an up-to-date expense', 'error')
            
            if not fields_to_update:
                flash('No edits were applied', 'error')
            else:
                db.property_expenses.update_one({'_id': ObjectId(item_id)},
                                    {'$set': fields_to_update})
                db.audit_logs.insert_one({'user': login_data,'Activity': 'Edit expense','Item': item_id,'timestamp': datetime.now()})
                flash('Expense updates were applied', 'success')
            return redirect('/view-property-expenses')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

####delete expense
@propertyManagement.route('/delete-property-expense/<item_id>', methods=['POST'])
def delete_property_expense(item_id):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            manager = db.registered_managers.find_one({'username':login_data},{'_id':0,'createdAt':0,'code':0,'address':0})
            if manager.get('add_properties') in ('yes', None):
                selected_item = db.property_expenses.find_one({'_id': ObjectId(item_id)})
                if selected_item:
                    db.property_expenses.delete_one({'_id': ObjectId(item_id)})
                    db.audit_logs.insert_one({'user': login_data,'Activity': 'Expense deletion','Item': item_id,'timestamp': datetime.now()})
                    flash('Expense was deleted', 'success')
                else:
                    flash('Expense does not exist', 'error')
            else:
                flash('You do not have rights to delete', 'error')
            return redirect('/view-property-expenses')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

###DOANLOAD EXPENSE DATA   
@propertyManagement.route('/download-property-expense-data', methods=["POST"])
def download_property_expense_data():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            startdate_on_str = request.form.get("startdate")
            enddate_on_str = request.form.get("enddate")
            startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
            enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')

            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'phone_number': 0, 'address': 0,
                                                                                'password': 0, 'auth': 0, 'dark_mode': 0})

            expenses = list(db.property_expenses.find(
                {'company_name': company['company_name'], 'expenseDate': {'$gte': startdate, '$lte': enddate}},
                {'_id': 0, 'company_name': 0}
            ))

            # Sort data by expenseDate in descending order
            sorted_expenses = sorted(expenses, key=lambda x: x["expenseDate"], reverse=True)

            # Create Excel file
            excel_buffer = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Expenses"

            # Write header row
            headers = ['Expense', 'Amount', 'Date', 'Property']
            ws.append(headers)

            # Write data rows
            for expense in sorted_expenses:
                row = [
                    expense.get('expenseName', ''),
                    expense.get('amount', 0),
                    expense.get('expenseDate', '').strftime('%Y-%m-%d') if isinstance(expense.get('expenseDate'), datetime) else '',
                    expense.get('propertyName', '')
                ]
                ws.append(row)

            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Create the response
            response = make_response(excel_buffer.getvalue())
            response.headers['Content-Disposition'] = f"attachment; filename={company['company_name']}_Expenses_{startdate_on_str}_{enddate_on_str}.xlsx"
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            # Clean up
            del wb
            del excel_buffer
            gc.collect()

            return response
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

############LOAD COMPLAINTS TO MANAGER######################
@propertyManagement.route('/resolve-complaints')
def resolve_complaints():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            if 'dp' in company:
                dp_str = company['dp']
            else:
                dp_str = None
            is_manager = db.managers.find_one({'manager_email': company['email'], 'name':company['company_name']})
            ####CHECK IS LOGEDIN MANAGER HAS FULL RIGHTS
            if is_manager is None:
                property_assigned = db.registered_managers.find({'username': login_data})
                property_assigned_dict = {property for doc in property_assigned if 'properties' in doc for property in doc['properties']}
                if not property_assigned_dict:
                    flash('You are not managing any property!', 'error')
                    return redirect('/load-dashboard-page')
                else:
                    tenant_accounts = []
                    for property in property_assigned_dict:
                        tenant_account = list(db.tenant_user_accounts.find({'propertyName': property}))
                        tenant_accounts.extend(tenant_account)
            else:
                user_querry = {'company_name': company['company_name']}
                properties = db.property_managed.find(user_querry)
                if db.property_managed.count_documents(user_querry)==0:
                    flash('You are not managing any property!', 'error')
                    return redirect('/load-dashboard-page')
                else:
                    tenant_accounts = []
                    for property in properties:
                        tenant_account = list(db.tenant_user_accounts.find({'propertyName': property['propertyName']}))
                        tenant_accounts.extend(tenant_account)

            complaints = []
            resolved_complaints=[]
            for tenant in tenant_accounts:
                tenant_id = tenant['_id']
                found_complaints = db.tenant_complaints.find({'tenantID': tenant_id})
                found_resolved_complaints = db.resolved_complaints.find({'tenantID': tenant_id})
                for found_resolved in found_resolved_complaints:
                    resolved_complaints.append(found_resolved)
                for complaint in found_complaints:
                    replies = list(db.tenant_complaints_replies.find({'complaintID': complaint['_id']}))
                    db.tenant_complaints.update_one({'_id': complaint['_id']}, {'$set': {'status': 'seen'}})
                    if len(replies) == 0:
                        replies = [{'Reply': 'No reply', 'who': 'N/A', 'reply_date': 'N/A'}]
                    else:
                        for reply in replies:
                            if reply['who'] != 'Manager':
                                db.tenant_complaints_replies.update_one({'_id': reply['_id']}, {'$set': {'status': 'seen'}})
                        # Sort replies by date, most recent first
                        replies = sorted(replies, key=lambda r: r['reply_date'], reverse=True)
                    complaint_copy = complaint.copy()  # create a copy of complaint to avoid overwriting
                    complaint_copy['_id'] = str(complaint['_id'])
                    complaint_copy['tenantID'] = str(complaint['tenantID'])
                    # Prepare replies with conditional status field
                    if len(replies) == 1 and replies[0]['Reply'] == 'No reply':
                        complaint_copy['replies'] = [{'Reply': replies[0]['Reply'], 'who': replies[0]['who'], 'reply_date': replies[0]['reply_date'], 'status': ''}]
                    else:
                        complaint_copy['replies'] = [{'Reply': reply['Reply'], 'who': reply['who'], 'other': 'Manager', 'status': reply.get('status', ''), 'reply_date': reply['reply_date'].strftime('%Y-%m-%d %H:%M') if reply['reply_date'] != 'N/A' else 'N/A'} for reply in replies]
                    complaints.append(complaint_copy)
            # Sort complaints by date, most recent first
            complaints = sorted(complaints, key=lambda c: c['complained_on'], reverse=True)
            # Remove duplicates
            complaints = list({v['_id']: v for v in complaints}.values())
            return render_template('resolve complaints.html',complaints=complaints,resolved_complaints=resolved_complaints, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
            
############RESOLVE COMPLAINTS BY MANAGER###########
@propertyManagement.route('/update-complaint', methods=['POST'])
def update_complaint():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            # manager = db.registered_managers.find_one({'username': login_data})
            send_emails = db.send_emails.find_one({'emails': "yes"},{'emails': 1})

            complaint_id = request.form.get('complaint_id')
            Reply = request.form.get('Reply_' + str(complaint_id))
            client_time_str = request.form.get('client_time')
            client_time = parse_iso_format(client_time_str)
            time_zone_offset = int(request.form.get('time_zone_offset'))
            adjusted_time = client_time + timedelta(hours=time_zone_offset)
            db.tenant_complaints_replies.insert_one({'complaintID': ObjectId(complaint_id),
                                                        'Reply': Reply,
                                                        'who': 'Manager',
                                                        'reply_date': adjusted_time,
                                                        'status': ''})
            
            tenant_complaint_id = db.tenant_complaints.find_one({'_id': ObjectId(complaint_id)})
            tenant_object_id = db.tenant_user_accounts.find_one({'_id': tenant_complaint_id['tenantID']})
            db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
            db.userNotifications.insert_one({
                'category': 'reply',
                'user': tenant_complaint_id['tenantID'],
                'notification': f"New reply from manager {login_data}",
                'timestamp': datetime.utcnow()
            })
            tenant_email = tenant_object_id['tenantEmail']

            if send_emails is not None:
                msg = Message('New Reply From Property Manager', 
                sender='michmanage@outlook.com', 
                recipients=[tenant_email])
                msg.html = f"""
                <html>
                <body>
                <p>Dear tenant,</p>
                <p>You have a new reply from manager of {tenant_object_id['propertyName']}, please login below to check reply:</p>
                <p><b style="font-size: 20px;"><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                <p>Best Regards,</p>
                <p>Mich Manage</p>
                </body>
                </html>
                """
                thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                thread.start()

            return redirect('/resolve-complaints')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
        
##########RESOLVING COMPLAINTS AFTER SOLVING THEM#########
@propertyManagement.route('/resolved-complaints/<complaint_id>', methods=["GET", "POST"])
def resolved_complaints(complaint_id):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            send_emails = db.send_emails.find_one({'emails': "yes"},{'emails': 1})
                
            resolved_complaint = db.tenant_complaints.find_one({'_id': ObjectId(complaint_id)})
            resolved_complaint['resolved_time'] = datetime.now()
            resolved_complaint['username'] = login_data
            db.resolved_complaints.insert_one(resolved_complaint)
            db.tenant_complaints.delete_one({'_id': ObjectId(complaint_id)})

            tenant = db.tenant_user_accounts.find_one({"_id": resolved_complaint["tenantID"]})
            tenant_email = tenant["tenantEmail"]
            manager = db.registered_managers.find_one({"username": login_data})
            manager_email = manager["email"]
            resolved_time = resolved_complaint["resolved_time"].replace(second=0, microsecond=0)
            complained_on = resolved_complaint["complained_on"].replace(second=0, microsecond=0)
            days_taken = ((resolved_time - complained_on).days) + 1

            db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
            db.userNotifications.insert_one({
                'category': 'reply',
                'user': resolved_complaint["tenantID"],
                'notification': f"Complaint resolved by manager {login_data}",
                'timestamp': datetime.utcnow()
            })

            if send_emails is not None:
                msg = Message('Complaint was resolved', 
                sender='michmanage@outlook.com', 
                recipients=[tenant_email, manager_email])
                msg.html = f"""
                <html>
                <body>
                <p>Dear user,</p>
                <p>The following complaint was resolved by {manager['name']}</p>
                <p>Heading: {resolved_complaint["complaint_heading"]}</p>
                <p>Details: {resolved_complaint["details"]}</p>
                <p>Date filed: {complained_on}</p>
                <p>Date resolved: {resolved_time}</p>
                <p>Time taken to resolve: {days_taken} days</p>
                <p>Best Regards,</p>
                <p>Mich Manage</p>
                </body>
                </html>
                """
                thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                thread.start()

            flash('Complaint was resolved', 'success')
            return redirect('/resolve-complaints')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
       
#############ADD PROPERTY####################
@propertyManagement.route('/add-property', methods=["POST"])
def add_property():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            all_items = request.json.get('items', [])  # Access the JSON data sent from the client
            skipped_items = []  # List to hold names of items that were not added
            added_items = []  # List to hold names of items that were successfully added
            for item in all_items:
                item['propertyName'] = item.get('propertyName', '').strip()
                item['sections'] = item.get('sections','').split(',')
                item['property_value'] = int(item['property_value'])
                item['late_payment_day'] = int(item['late_payment_day'])

                property_exists = db.property_managed.find_one({'propertyName': item['propertyName']})
                manager = db.registered_managers.find_one({'username':login_data})
                
                if property_exists is None:
                    item['username'] = login_data
                    item['company_name'] = manager['company_name']
                    db.property_managed.insert_one(item)
                    db.audit_logs.insert_one({'user': login_data, 'Activity': 'Add property data', 'propertyName':item['propertyName'], 'timestamp': datetime.now()})
                    added_items.append(item['propertyName'])
                else:
                    skipped_items.append(item.get('propertyName', 'unknown'))

            message = ""
            if added_items:
                message += 'The following properties were added: ' + ', '.join(added_items)
                flash(message, 'success')
            if skipped_items:
                message_skipped = 'The following properties were not added because they already exist: ' + ', '.join(skipped_items)
                flash(message_skipped, 'error')
            return jsonify({'redirect': url_for('propertyManagement_route.load_dashboard_page')})
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

########LOAD TENANT INFO################
@propertyManagement.route('/update-tenant-info')
def update_tenant_info():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    current_year = datetime.now().year
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            dp_str = None

            is_manager = db.managers.find_one({'manager_email': company['email'], 'name':company['company_name']}) is not None

            if not is_manager:
                property_assigned = db.registered_managers.find({'username': login_data})
                property_assigned_dict = {property for doc in property_assigned if 'properties' in doc for property in doc['properties']}
                tenants = []
                for property in property_assigned_dict:
                    properties_query = {"propertyName": property}
                    tenants_data = list(db.tenants.find(properties_query))
                    if tenants_data:
                        for tenant in tenants_data:
                            tenants.append(tenant)
                property_managed = []
                for property in property_assigned_dict:
                    properties_query = {"propertyName": property}
                    property_data = list(db.property_managed.find(properties_query))
                    if property_data:
                        for property in property_data:
                            property_managed.append(property)
            else:
                tenants_query = {'company_name': company['company_name']}
                property_query = {'company_name': company['company_name']}
                tenants = list(db.tenants.find(tenants_query))
                property_managed = list(db.property_managed.find(property_query))
            
            if not tenants:
                flash('No tenant data found', 'error')
                return redirect('/load-dashboard-page')
            
            tenant_data = []
            month_mapping = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
                'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12,
                'Quarter 1': 3, 'Quarter 2': 6, 'Quarter 3': 9, 'Quarter 4': 12,
                '2024': 12, '2025': 12, '2026': 12
            }
            for tenant in tenants:
                for property in property_managed:
                    if tenant['propertyName'] == property['propertyName']:
                        last_payment_month = month_mapping.get(tenant['months_paid'], 0)
                        last_payment_date = datetime(year=tenant["date_last_paid"].year, month=last_payment_month, day=1)
                        next_payment_date = last_payment_date + timedelta(days=30)
                        remaining_days = (next_payment_date - datetime.now()).days
                        percentage_paid = round((tenant['available_amount'] / tenant['section_value']) * 100, 1)
                        date_last_paid = tenant['date_last_paid'].strftime('%Y-%m-%d')
                        amount_demanded = max(0, tenant['section_value'] - tenant['available_amount'])
                        if remaining_days < 0:
                            overdue = True
                            remaining_days = abs(remaining_days)
                            amount_next_month = int((round((remaining_days) / 30 + 0.5, 0)) * tenant['section_value'])
                            amount_demanded = (tenant['section_value'] - tenant['available_amount']) + amount_next_month                   
                        else:
                            overdue = False

                        if remaining_days < 7:
                            time_unit = 'day(s)'
                        elif remaining_days < 30:
                            remaining_days = round(remaining_days / 7)
                            time_unit = 'week(s)'
                        elif remaining_days < 365:
                            remaining_days = round(remaining_days / 30)
                            time_unit = 'month(s)'
                        else:
                            remaining_days = round(remaining_days / 365)
                            time_unit = 'year(s)'

                        overdue_status = 'overdue' if overdue else 'due in'
                        tenant_data.append((tenant['_id'],tenant['tenantName'], tenant['tenantPhone'], tenant['tenantEmail'],
                                            tenant['propertyName'], tenant['selected_section'], tenant['payment_type'],
                                            tenant['amount'], tenant['payment_mode'], tenant['months_paid'], 
                                            tenant['available_amount'], amount_demanded, tenant['payment_completion'], 
                                            tenant['section_value'], percentage_paid, tenant['currency'], 
                                            date_last_paid, remaining_days, time_unit, overdue_status))
            
            if 'dp' in company:
                dp = base64.b64decode(company['dp'])
                dp_str = base64.b64encode(dp).decode()
            return render_template('tenant information.html', tenant_data=tenant_data, dp=dp_str, current_year=current_year)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@propertyManagement.route('/get_receipt', methods=['GET'])
def get_receipt():
    db, fs = get_db_and_fs()
    tenant_email = request.args.get('tenantEmail')
    property_name = request.args.get('propertyName')
    selected_section = request.args.get('selected_section')
    months_paid = request.args.get('months_paid')
    year = request.args.get('year')
    year = int(year)

    receipt_data = db.tenants.find_one({
        'tenantEmail': tenant_email,
        'propertyName': property_name,
        'selected_section': selected_section,
        'months_paid': months_paid,
        'year': year
    }, {'payment_receipt': 1, '_id': 0})

    if receipt_data is None:
        old_receipt_data = db.old_tenant_data.find_one({
            'tenantEmail': tenant_email,
            'propertyName': property_name,
            'selected_section': selected_section,
            'months_paid': months_paid,
            'year': year
        }, {'payment_receipt': 1, '_id': 0})
        if old_receipt_data:
            # Convert the base64 string back to bytes
            payment_receipt = base64.b64decode(old_receipt_data['payment_receipt'])

            # Create a BytesIO object from the PDF data
            pdf_io = io.BytesIO(payment_receipt)

            # Create the file name
            file_name = f"{property_name}_{selected_section}_{months_paid}_{year}.pdf"

            # Create a custom response
            response = make_response(pdf_io.getvalue())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename={file_name}'

            return response

    else:
        # Convert the base64 string back to bytes
        payment_receipt = base64.b64decode(receipt_data['payment_receipt'])

        # Create a BytesIO object from the PDF data
        pdf_io = io.BytesIO(payment_receipt)

        # Create the file name
        file_name = f"{property_name}_{selected_section}_{months_paid}_{year}.pdf"

        # Create a custom response
        response = make_response(pdf_io.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={file_name}'

        return response

###########UPDATE TENANT INFO################
@propertyManagement.route('/update', methods=['POST'])
def update():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            send_emails = db.send_emails.find_one({'emails': "yes"},{'emails': 1})
            
            tenantId = request.form.get('tenantId')
            section_tenant = db.tenants.find_one({'_id': ObjectId(tenantId)})

            new_amount_from_form = request.form.get('amount_paid')
            payment_mode = request.form.get('payment_mode')
            months_paid = request.form.get('months_paid')
            date = request.form.get('date')
            tenantEmail = section_tenant['tenantEmail']
            propertyName = section_tenant['propertyName']
            selected_section = section_tenant['selected_section']
            new_amount = int(new_amount_from_form)
            # Convert the date string to a datetime object
            date = datetime.strptime(date, '%Y-%m-%d')

            section_value = section_tenant['section_value']
            payment_status = ""

            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})

            old_data = db.tenants.find_one({'tenantEmail': tenantEmail, 'propertyName':propertyName, 'selected_section': selected_section})
                    
            old_amount = old_data['available_amount']
            old_date = old_data['date_last_paid']
            late_payment_day = (db.property_managed.find_one({'propertyName': propertyName}))['late_payment_day']
            if date.day > late_payment_day:
                payment_status = "Late"
            else:
                payment_status = "Early"
            payment_completion = ''
            month_mapping = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
                'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12,
                'Quarter 1': 3, 'Quarter 2': 6, 'Quarter 3': 9, 'Quarter 4': 12,
                '2024': 12, '2025': 12, '2026': 12
            }

            available_amount = 0
            field_month = month_mapping.get(old_data['months_paid'], 0)
            months_paid_selected = month_mapping.get(months_paid, 0)
            if field_month != months_paid_selected:
                # Define the start and end of the current year
                start_date = datetime(date.year, 1, 1)
                end_date = datetime(date.year + 1, 1, 1)

                old_data2 = db.old_tenant_data.find_one({'tenantEmail': tenantEmail, 'propertyName':propertyName, 'selected_section': selected_section, 'months_paid': months_paid, 'date_last_paid': {'$gte': start_date,'$lt': end_date}})

                if old_data2:
                    flash('Selected period was fully paid', 'error')
                else:
                    if old_data['available_amount'] < section_value:
                        flash('First fully update current/previous period', 'error')
                    else:
                        available_amount = new_amount

                        balance = section_value - available_amount
                        # Create a payment receipt PDF file
                        buffer = BytesIO()
                        doc = SimpleDocTemplate(buffer, pagesize=letter)

                        # QR Code Generation
                        url = f'https://michmanagement.onrender.com//get_receipt?tenantEmail={tenantEmail}&propertyName={propertyName}&selected_section={selected_section}&months_paid={months_paid}&year={date.year}'
                        qr = qrcode.QRCode(
                            version=1,
                            error_correction=qrcode.constants.ERROR_CORRECT_L,
                            box_size=3,
                            border=4,
                        )
                        qr.add_data(url)
                        qr.make(fit=True)
                        img = qr.make_image(fill_color="black", back_color="white")
                        login_data = session.get('login_username')
                        img.save(f'payment_receipt_qr_{login_data}.png')

                        # Create the receipt details
                        data = [
                            ['Rent Payment Receipt - ' + company['company_name'], ''],
                            ['Receipt for:', section_tenant['tenantName']],
                            ['Property Name:', propertyName],
                            ['Payment Type:', section_tenant['payment_type']],
                            ['Amount Paid:', f"{section_tenant['currency']} {new_amount}"],
                            ['Payment Mode:', payment_mode],
                            ['Month Paid for:', months_paid],
                            ['Date Paid:', date.strftime('%Y-%m-%d')],
                            ['Balance:', f"{section_tenant['currency']} {balance}"],
                            ['Prepared by:', company['name']]
                        ]

                        # Create a table with the receipt details
                        table = Table(data)

                        # Add a table style
                        table.setStyle(TableStyle([
                            ('SPAN', (0, 0), (1, 0)),  # Merge the first row
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),

                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 14),

                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0,0), (-1,-1), 1, colors.black),
                            ('FONTNAME', (1, -1), (1, -1), 'Helvetica-Oblique')  # Make the last cell on the last row italic
                        ]))

                        # Load your QR code image
                        qr_code_img = f'payment_receipt_qr_{login_data}.png'
                        qr_code = Image(qr_code_img)
                        qr_code.hAlign = 'CENTER'

                        # Add the QR code image to the elements list before building the PDF
                        elements = [table, qr_code]
                        doc.build(elements)

                        # Get the PDF data and encode it as base64
                        pdf_data = buffer.getvalue()
                        buffer.close()
                        payment_receipt_base64 = base64.b64encode(pdf_data).decode()

                        # Delete the QR code image file
                        os.remove(f'payment_receipt_qr_{login_data}.png')
                        
                        new_data = {
                            'username': login_data,
                            'company_name': company['company_name'],
                            'tenantName': old_data['tenantName'],
                            'tenantEmail': tenantEmail,
                            'tenantPhone': old_data['tenantPhone'],
                            'propertyName': propertyName,
                            'selected_section': selected_section,
                            'section_value': old_data['section_value'],
                            'payment_type': old_data['payment_type'],
                            'amount': new_amount,
                            'payment_mode': payment_mode,
                            'months_paid': months_paid,
                            'year': date.year,
                            'available_amount': available_amount,
                            'payment_completion': payment_completion,
                            'currency': old_data['currency'],
                            'date_last_paid': date,
                            'payment_status': payment_status,
                            'status': 'updated',
                            'payment_receipt': payment_receipt_base64
                        }
                        if available_amount < section_value:
                            payment_completion = 'Partial'
                            new_data['payment_completion'] = payment_completion
                            if date.year == old_date.year:
                                if field_month > months_paid_selected:
                                    db.old_tenant_data.insert_one(new_data)
                                    tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                                    if tenant_user:
                                        db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                        db.userNotifications.insert_one({
                                            'category': 'payment',
                                            'user': tenant_user["_id"],
                                            'notification': f"New payment recorded by manager {login_data}",
                                            'timestamp': datetime.utcnow()
                                        })
                                    # Create the email message
                                    if send_emails is not None:
                                        msg = Message('Rent Payment Receipt-Mich Manage', 
                                                    sender='michmanage@outlook.com', 
                                                    recipients=[tenantEmail])
                                        msg.html = f"""
                                        <html>
                                        <body>
                                        <p>Dear {section_tenant['tenantName']},</p>
                                        <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                        <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                        <p>Best Regards,</p>
                                        <p>Mich Manage</p>
                                        </body>
                                        </html>
                                        """

                                        # Attach the PDF receipt to the email
                                        msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                        # Send the email
                                        thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                        thread.start()
                                    db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantName': old_data['tenantName'], 'timestamp': datetime.now()})
                                    flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                                else:
                                    db.tenants.update_one({'_id': ObjectId(old_data['_id'])}, {'$set': new_data})
                                    tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                                    if tenant_user:
                                        db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                        db.userNotifications.insert_one({
                                            'category': 'payment',
                                            'user': tenant_user["_id"],
                                            'notification': f"New payment recorded by manager {login_data}",
                                            'timestamp': datetime.utcnow()
                                        })
                                    # Create the email message
                                    if send_emails is not None:
                                        msg = Message('Rent Payment Receipt-Mich Manage', 
                                                    sender='michmanage@outlook.com', 
                                                    recipients=[tenantEmail])
                                        msg.html = f"""
                                        <html>
                                        <body>
                                        <p>Dear {section_tenant['tenantName']},</p>
                                        <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                        <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                        <p>Best Regards,</p>
                                        <p>Mich Manage</p>
                                        </body>
                                        </html>
                                        """

                                        # Attach the PDF receipt to the email
                                        msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                        # Send the email
                                        thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                        thread.start()

                                    if '_id' in old_data:
                                        del old_data['_id']
                                    db.old_tenant_data.insert_one(old_data)
                                    db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantName': old_data['tenantName'], 'timestamp': datetime.now()})
                                    flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                            elif date.year < old_date.year:
                                db.old_tenant_data.insert_one(new_data)
                                tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                                if tenant_user:
                                    db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                    db.userNotifications.insert_one({
                                        'category': 'payment',
                                        'user': tenant_user["_id"],
                                        'notification': f"New payment recorded by manager {login_data}",
                                        'timestamp': datetime.utcnow()
                                    })
                                # Create the email message
                                if send_emails is not None:
                                    msg = Message('Rent Payment Receipt-Mich Manage', 
                                                sender='michmanage@outlook.com', 
                                                recipients=[tenantEmail])
                                    msg.html = f"""
                                    <html>
                                    <body>
                                    <p>Dear {section_tenant['tenantName']},</p>
                                    <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                    <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                    <p>Best Regards,</p>
                                    <p>Mich Manage</p>
                                    </body>
                                    </html>
                                    """

                                    # Attach the PDF receipt to the email
                                    msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                    # Send the email
                                    thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                    thread.start()
                                db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantName': old_data['tenantName'], 'timestamp': datetime.now()})
                                flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                            else:
                                db.tenants.update_one({'_id': ObjectId(old_data['_id'])}, {'$set': new_data})
                                tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                                if tenant_user:
                                    db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                    db.userNotifications.insert_one({
                                        'category': 'payment',
                                        'user': tenant_user["_id"],
                                        'notification': f"New payment recorded by manager {login_data}",
                                        'timestamp': datetime.utcnow()
                                    })
                                # Create the email message
                                if send_emails is not None:
                                    msg = Message('Rent Payment Receipt-Mich Manage', 
                                                sender='michmanage@outlook.com', 
                                                recipients=[tenantEmail])
                                    msg.html = f"""
                                    <html>
                                    <body>
                                    <p>Dear {section_tenant['tenantName']},</p>
                                    <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                    <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                    <p>Best Regards,</p>
                                    <p>Mich Manage</p>
                                    </body>
                                    </html>
                                    """

                                    # Attach the PDF receipt to the email
                                    msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                    # Send the email
                                    thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                    thread.start()
                                if '_id' in old_data:
                                    del old_data['_id']
                                db.old_tenant_data.insert_one(old_data)
                                db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantName': old_data['tenantName'], 'timestamp': datetime.now()})
                                flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                        elif available_amount > section_value:  
                            flash("Enter amount that does not exceed section value", 'error')

                        else:
                            payment_completion = 'Full'
                            new_data['payment_completion'] = payment_completion

                            if date.year == old_date.year:
                                if field_month > months_paid_selected:
                                    db.old_tenant_data.insert_one(new_data)
                                    tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                                    if tenant_user:
                                        db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                        db.userNotifications.insert_one({
                                            'category': 'payment',
                                            'user': tenant_user["_id"],
                                            'notification': f"New payment recorded by manager {login_data}",
                                            'timestamp': datetime.utcnow()
                                        })
                                    # Create the email message
                                    if send_emails is not None:
                                        msg = Message('Rent Payment Receipt-Mich Manage', 
                                                    sender='michmanage@outlook.com', 
                                                    recipients=[tenantEmail])
                                        msg.html = f"""
                                        <html>
                                        <body>
                                        <p>Dear {section_tenant['tenantName']},</p>
                                        <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                        <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                        <p>Best Regards,</p>
                                        <p>Mich Manage</p>
                                        </body>
                                        </html>
                                        """

                                        # Attach the PDF receipt to the email
                                        msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                        # Send the email
                                        thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                        thread.start()
                                    db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantName': old_data['tenantName'], 'timestamp': datetime.now()})
                                    flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                                else:
                                    db.tenants.update_one({'_id': ObjectId(old_data['_id'])}, {'$set': new_data})
                                    tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                                    if tenant_user:
                                        db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                        db.userNotifications.insert_one({
                                            'category': 'payment',
                                            'user': tenant_user["_id"],
                                            'notification': f"New payment recorded by manager {login_data}",
                                            'timestamp': datetime.utcnow()
                                        })
                                    # Create the email message
                                    if send_emails is not None:
                                        msg = Message('Rent Payment Receipt-Mich Manage', 
                                                    sender='michmanage@outlook.com', 
                                                    recipients=[tenantEmail])
                                        msg.html = f"""
                                        <html>
                                        <body>
                                        <p>Dear {section_tenant['tenantName']},</p>
                                        <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                        <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                        <p>Best Regards,</p>
                                        <p>Mich Manage</p>
                                        </body>
                                        </html>
                                        """

                                        # Attach the PDF receipt to the email
                                        msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                        # Send the email
                                        thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                        thread.start()
                                    if '_id' in old_data:
                                        del old_data['_id']
                                    db.old_tenant_data.insert_one(old_data)
                                    db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantName': old_data['tenantName'], 'timestamp': datetime.now()})
                                    flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                            elif date.year < old_date.year:
                                db.old_tenant_data.insert_one(new_data)
                                tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                                if tenant_user:
                                    db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                    db.userNotifications.insert_one({
                                        'category': 'payment',
                                        'user': tenant_user["_id"],
                                        'notification': f"New payment recorded by manager {login_data}",
                                        'timestamp': datetime.utcnow()
                                    })
                                # Create the email message
                                if send_emails is not None:
                                    msg = Message('Rent Payment Receipt-Mich Manage', 
                                                sender='michmanage@outlook.com', 
                                                recipients=[tenantEmail])
                                    msg.html = f"""
                                    <html>
                                    <body>
                                    <p>Dear {section_tenant['tenantName']},</p>
                                    <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                    <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                    <p>Best Regards,</p>
                                    <p>Mich Manage</p>
                                    </body>
                                    </html>
                                    """

                                    # Attach the PDF receipt to the email
                                    msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                    # Send the email
                                    thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                    thread.start()
                                db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantName': old_data['tenantName'], 'timestamp': datetime.now()})
                                flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                            else:
                                db.tenants.update_one({'_id': ObjectId(old_data['_id'])}, {'$set': new_data})
                                tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                                if tenant_user:
                                    db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                    db.userNotifications.insert_one({
                                        'category': 'payment',
                                        'user': tenant_user["_id"],
                                        'notification': f"New payment recorded by manager {login_data}",
                                        'timestamp': datetime.utcnow()
                                    })
                                # Create the email message
                                if send_emails is not None:
                                    msg = Message('Rent Payment Receipt-Mich Manage', 
                                                sender='michmanage@outlook.com', 
                                                recipients=[tenantEmail])
                                    msg.html = f"""
                                    <html>
                                    <body>
                                    <p>Dear {section_tenant['tenantName']},</p>
                                    <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                    <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                    <p>Best Regards,</p>
                                    <p>Mich Manage</p>
                                    </body>
                                    </html>
                                    """

                                    # Attach the PDF receipt to the email
                                    msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                    # Send the email
                                    thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                    thread.start()
                                if '_id' in old_data:
                                    del old_data['_id']
                                db.old_tenant_data.insert_one(old_data)
                                db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantName': old_data['tenantName'], 'timestamp': datetime.now()})
                                flash(f"Updates for {old_data['tenantName']} were successful", 'success')

            elif field_month == months_paid_selected:
                if old_data['available_amount'] == section_value:
                    flash('Period selected is fully paid', 'error')
                else:
                    available_amount = new_amount + old_amount

                    balance = section_value - available_amount

                    buffer = BytesIO()
                    doc = SimpleDocTemplate(buffer, pagesize=letter)

                    # QR Code Generation
                    url = f'https://michmanagement.onrender.com//get_receipt?tenantEmail={tenantEmail}&propertyName={propertyName}&selected_section={selected_section}&months_paid={months_paid}&year={date.year}'
                    qr = qrcode.QRCode(
                        version=1,
                        error_correction=qrcode.constants.ERROR_CORRECT_L,
                        box_size=3,
                        border=4,
                    )
                    qr.add_data(url)
                    qr.make(fit=True)
                    img = qr.make_image(fill_color="black", back_color="white")
                    img.save(f'payment_receipt_qr_{login_data}.png')

                    # Create the receipt details
                    data = [
                        ['Rent Payment Receipt - ' + company['company_name'], ''],
                        ['Receipt for:', section_tenant['tenantName']],
                        ['Property Name:', propertyName],
                        ['Payment Type:', section_tenant['payment_type']],
                        ['Amount Paid:', f"{section_tenant['currency']} {new_amount}"],
                        ['Payment Mode:', payment_mode],
                        ['Month Paid for:', months_paid],
                        ['Date Paid:', date.strftime('%Y-%m-%d')],
                        ['Balance:', f"{section_tenant['currency']} {balance}"],
                        ['Prepared by:', company['name']]
                    ]

                    # Create a table with the receipt details
                    table = Table(data)

                    # Add a table style
                    table.setStyle(TableStyle([
                        ('SPAN', (0, 0), (1, 0)),  # Merge the first row
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),

                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 14),

                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0,0), (-1,-1), 1, colors.black),
                        ('FONTNAME', (1, -1), (1, -1), 'Helvetica-Oblique')  # Make the last cell on the last row italic
                    ]))

                    # Load your QR code image
                    qr_code_img = f'payment_receipt_qr_{login_data}.png'
                    qr_code = Image(qr_code_img)
                    qr_code.hAlign = 'CENTER'

                    # Add the QR code image to the elements list before building the PDF
                    elements = [table, qr_code]
                    doc.build(elements)

                    # Get the PDF data and encode it as base64
                    pdf_data = buffer.getvalue()
                    buffer.close()
                    payment_receipt_base64 = base64.b64encode(pdf_data).decode()

                    # Delete the QR code image file
                    os.remove(f'payment_receipt_qr_{login_data}.png')
                    
                    new_data = {
                        'username': login_data,
                        'company_name': company['company_name'],
                        'tenantName': old_data['tenantName'],
                        'tenantEmail': tenantEmail,
                        'tenantPhone': old_data['tenantPhone'],
                        'propertyName': propertyName,
                        'selected_section': selected_section,
                        'section_value': old_data['section_value'],
                        'payment_type': old_data['payment_type'],
                        'amount': new_amount,
                        'payment_mode': payment_mode,
                        'months_paid': months_paid,
                        'year': date.year,
                        'available_amount': available_amount,
                        'payment_completion': payment_completion,
                        'currency': old_data['currency'],
                        'date_last_paid': date,
                        'payment_status': payment_status,
                        'status': 'updated',
                        'payment_receipt': payment_receipt_base64
                    }
                    if available_amount < section_value:
                        payment_completion = 'Partial'
                        new_data['payment_completion'] = payment_completion
                        if date.year == old_date.year:
                            db.tenants.update_one({'_id': ObjectId(old_data['_id'])}, {'$set': new_data})
                            tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                            if tenant_user:
                                db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                db.userNotifications.insert_one({
                                    'category': 'payment',
                                    'user': tenant_user["_id"],
                                    'notification': f"New payment recorded by manager {login_data}",
                                    'timestamp': datetime.utcnow()
                                })
                            # Create the email message
                            if send_emails is not None:
                                msg = Message('Rent Payment Receipt-Mich Manage', 
                                            sender='michmanage@outlook.com', 
                                            recipients=[tenantEmail])
                                msg.html = f"""
                                <html>
                                <body>
                                <p>Dear {section_tenant['tenantName']},</p>
                                <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                <p>Best Regards,</p>
                                <p>Mich Manage</p>
                                </body>
                                </html>
                                """

                                # Attach the PDF receipt to the email
                                msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                # Send the email
                                thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                thread.start()
                            db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantEmail':tenantEmail, 'timestamp': datetime.now()})
                            flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                        elif date.year < old_date.year:
                            db.old_tenant_data.insert_one(new_data)
                            tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                            if tenant_user:
                                db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                db.userNotifications.insert_one({
                                    'category': 'payment',
                                    'user': tenant_user["_id"],
                                    'notification': f"New payment recorded by manager {login_data}",
                                    'timestamp': datetime.utcnow()
                                })
                            # Create the email message
                            if send_emails is not None:
                                msg = Message('Rent Payment Receipt-Mich Manage', 
                                            sender='michmanage@outlook.com', 
                                            recipients=[tenantEmail])
                                msg.html = f"""
                                <html>
                                <body>
                                <p>Dear {section_tenant['tenantName']},</p>
                                <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                <p>Best Regards,</p>
                                <p>Mich Manage</p>
                                </body>
                                </html>
                                """

                                # Attach the PDF receipt to the email
                                msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                # Send the email
                                thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                thread.start()
                            db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantEmail':tenantEmail, 'timestamp': datetime.now()})
                            flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                        else:
                            db.tenants.update_one({'_id': ObjectId(old_data['_id'])}, {'$set': new_data})
                            tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                            if tenant_user:
                                db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                db.userNotifications.insert_one({
                                    'category': 'payment',
                                    'user': tenant_user["_id"],
                                    'notification': f"New payment recorded by manager {login_data}",
                                    'timestamp': datetime.utcnow()
                                })
                            # Create the email message
                            if send_emails is not None:
                                msg = Message('Rent Payment Receipt-Mich Manage', 
                                            sender='michmanage@outlook.com', 
                                            recipients=[tenantEmail])
                                msg.html = f"""
                                <html>
                                <body>
                                <p>Dear {section_tenant['tenantName']},</p>
                                <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                <p>Best Regards,</p>
                                <p>Mich Manage</p>
                                </body>
                                </html>
                                """

                                # Attach the PDF receipt to the email
                                msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                # Send the email
                                thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                thread.start()
                            if '_id' in old_data:
                                del old_data['_id']
                            db.old_tenant_data.insert_one(old_data)
                            db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantEmail':tenantEmail, 'timestamp': datetime.now()})
                            flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                    elif available_amount > section_value:        
                        flash("Enter amount that does not exceed section value", 'error')
                    else:
                        payment_completion = 'Full'
                        new_data['payment_completion'] = payment_completion
                        if date.year == old_date.year:
                            db.tenants.update_one({'_id': ObjectId(old_data['_id'])}, {'$set': new_data})
                            tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                            if tenant_user:
                                db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                db.userNotifications.insert_one({
                                    'category': 'payment',
                                    'user': tenant_user["_id"],
                                    'notification': f"New payment recorded by manager {login_data}",
                                    'timestamp': datetime.utcnow()
                                })
                            # Create the email message
                            if send_emails is not None:
                                msg = Message('Rent Payment Receipt-Mich Manage', 
                                            sender='michmanage@outlook.com', 
                                            recipients=[tenantEmail])
                                msg.html = f"""
                                <html>
                                <body>
                                <p>Dear {section_tenant['tenantName']},</p>
                                <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                <p>Best Regards,</p>
                                <p>Mich Manage</p>
                                </body>
                                </html>
                                """

                                # Attach the PDF receipt to the email
                                msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                # Send the email
                                thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                thread.start()
                            db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantEmail':tenantEmail, 'timestamp': datetime.now()})
                            flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                        elif date.year < old_date.year:
                            db.old_tenant_data.insert_one(new_data)
                            tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                            if tenant_user:
                                db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                db.userNotifications.insert_one({
                                    'category': 'payment',
                                    'user': tenant_user["_id"],
                                    'notification': f"New payment recorded by manager {login_data}",
                                    'timestamp': datetime.utcnow()
                                })
                            # Create the email message
                            if send_emails is not None:
                                msg = Message('Rent Payment Receipt-Mich Manage', 
                                            sender='michmanage@outlook.com', 
                                            recipients=[tenantEmail])
                                msg.html = f"""
                                <html>
                                <body>
                                <p>Dear {section_tenant['tenantName']},</p>
                                <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                <p>Best Regards,</p>
                                <p>Mich Manage</p>
                                </body>
                                </html>
                                """

                                # Attach the PDF receipt to the email
                                msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                # Send the email
                                thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                thread.start()
                            db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantEmail':tenantEmail, 'timestamp': datetime.now()})
                            flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                        else:
                            db.tenants.update_one({'_id': ObjectId(old_data['_id'])}, {'$set': new_data})
                            tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                            if tenant_user:
                                db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                                db.userNotifications.insert_one({
                                    'category': 'payment',
                                    'user': tenant_user["_id"],
                                    'notification': f"New payment recorded by manager {login_data}",
                                    'timestamp': datetime.utcnow()
                                })
                            # Create the email message
                            if send_emails is not None:
                                msg = Message('Rent Payment Receipt-Mich Manage', 
                                            sender='michmanage@outlook.com', 
                                            recipients=[tenantEmail])
                                msg.html = f"""
                                <html>
                                <body>
                                <p>Dear {section_tenant['tenantName']},</p>
                                <p>Please find attached your payment receipt for {months_paid} {date.year}.</p>
                                <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                                <p>Best Regards,</p>
                                <p>Mich Manage</p>
                                </body>
                                </html>
                                """

                                # Attach the PDF receipt to the email
                                msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                                # Send the email
                                thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                                thread.start()
                            if '_id' in old_data:
                                del old_data['_id']
                            db.old_tenant_data.insert_one(old_data)
                            db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update tenant data', 'tenantEmail':tenantEmail, 'timestamp': datetime.now()})
                            flash(f"Updates for {old_data['tenantName']} were successful", 'success')
                    
            return redirect('/update-tenant-info')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

########LOAD PROPERTY DATA ################
def get_property_data(properties):
    property_data = []
    for property in properties:
        property_data.append((property['propertyName'], property['type'], property['property_value'],
                              property['address'], property['owner_name'], property['owner_phone']))
    return property_data

@propertyManagement.route('/view-property-info')
def view_property_info():
    db, fs = get_db_and_fs()
    username = session.get('login_username')
    if username is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': username})
            if 'dp' in company:
                dp_str = company['dp']
            else:
                dp_str = None
            is_manager = db.managers.find_one({'manager_email': company['email'], 'name':company['company_name']}) is not None

            properties_query = {'company_name': company['company_name']}
            if not is_manager:
                property_assigned = db.registered_managers.find({'username': username})
                property_assigned_dict = {property for doc in property_assigned if 'properties' in doc for property in doc['properties']}
                property_data = []
                for property in property_assigned_dict:
                    properties_query = {"propertyName": property}
                    properties = list(db.property_managed.find(properties_query))
                    if properties:
                        property_data.extend(get_property_data(properties))
                return render_template('property information.html', property_data=property_data, dp=dp_str)
            else:
                properties = list(db.property_managed.find(properties_query))
                if not properties:
                    flash('We did not find property data', 'error')
                    return redirect('/load-dashboard-page')
                
                property_data = get_property_data(properties)
                return render_template('property information.html', property_data=property_data, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

#####UPDATE PROPERTY INFO#############
@propertyManagement.route('/update-property/<propertyName>')
def selected_property(propertyName):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            if 'dp' in company:
                dp_str = company['dp']
            else:
                dp_str = None
            return render_template('update property information.html',propertyName=propertyName, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

##POSTING NEW PROPERTY INFORMATION
@propertyManagement.route('/update-property', methods=["POST"])
def update_property():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            propertyName = request.form.get("propertyName")
            property_value = request.form.get("property_value")
            owner_name = request.form.get("owner_name")
            owner_email = request.form.get("owner_email")
            owner_phone = request.form.get("owner_phone")
            owner_residence = request.form.get("owner_residence")
            

            update_fields = {}

            if property_value:
                property_value = int(property_value)
                update_fields['property_value'] = property_value
            if owner_name:
                update_fields['owner_name'] = owner_name
            if owner_email:
                update_fields['owner_email'] = owner_email
            if owner_phone:
                update_fields['owner_phone'] = owner_phone
            if owner_residence:
                update_fields['owner_residence'] = owner_residence

            # Update the document with the non-empty fields
            db.property_managed.update_one({'propertyName': propertyName}, {'$set': update_fields})
            flash(f"{propertyName} was successfully updated", 'success')
            return redirect('/view-property-info')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

#######CLICK TO UPDATE TENANT#############
@propertyManagement.route('/selected-tenant/<tenantId>')
def selected_tenant(tenantId):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error') 
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            non_manager_update = session.get('update_tenant')
            if non_manager_update is None:
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                tenant = db.tenants.find_one({'_id': ObjectId(tenantId)})
                date_last_paid = tenant['date_last_paid']
                return render_template('update tenant information.html',tenantId=tenantId,tenantName=tenant['tenantName'],propertyName=tenant['propertyName'],selected_section=tenant['selected_section'],payment_type=tenant['payment_type'],amount=tenant['amount'],months_paid=tenant['months_paid'],year=date_last_paid.year,dp=dp_str)
            else:
                if non_manager_update == "yes":
                    company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                    if 'dp' in company:
                        dp_str = company['dp']
                    else:
                        dp_str = None
                    tenant = db.tenants.find_one({'_id': ObjectId(tenantId)})
                    date_last_paid = tenant['date_last_paid']
                    return render_template('update tenant information.html',tenantId=tenantId,tenantName=tenant['tenantName'],propertyName=tenant['propertyName'],selected_section=tenant['selected_section'],payment_type=tenant['payment_type'],amount=tenant['amount'],months_paid=tenant['months_paid'],year=date_last_paid.year,dp=dp_str)
                else:
                    flash('You do not have rights to update tenants', 'error')
                    return redirect('/update-tenant-info')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
        
##########EDIT TENANT INFO###################
@propertyManagement.route('/edit/<tenantId>')
def edit(tenantId):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            non_manager_update = session.get('edit_tenant')
            if non_manager_update is None:
                # Retrieve the tenant's info using the email
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                tenant = db.tenants.find_one({'_id': ObjectId(tenantId)})
                if tenant is None:
                    return "Tenant not found", 404
                # Pass the tenant's info to the template
                return render_template('edit.html',tenantId=tenantId,tenantName=tenant['tenantName'], payment_type=tenant['payment_type'], dp=dp_str)
            else:
                if non_manager_update == "yes":
                    # Retrieve the tenant's info using the email
                    company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                    if 'dp' in company:
                        dp_str = company['dp']
                    else:
                        dp_str = None
                    tenant = db.tenants.find_one({'_id': ObjectId(tenantId)})
                    if tenant is None:
                        return "Tenant not found", 404
                    # Pass the tenant's info to the template
                    return render_template('edit.html',tenantId=tenantId,tenantName=tenant['tenantName'], payment_type=tenant['payment_type'], dp=dp_str)
                else:
                    flash('You do not have rights to edit tenant information', 'error')
                    return redirect('/update-tenant-info')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

############APPLY EDITS##############
@propertyManagement.route('/make-edits', methods=["POST"])
def make_edits():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            send_emails = db.send_emails.find_one({'emails': "yes"},{'emails': 1})

            tenantId = request.form.get('tenantId')
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            # Create a dictionary for the fields to update
            fields_to_update = {}
            fields_to_update['status'] = 'edited'
            section_value = request.form.get('section_value')
            tenant = db.tenants.find_one({'_id': ObjectId(tenantId)})
            if section_value:
                section_value = int(section_value)
                fields_to_update['section_value'] = section_value
            else:
                section_value = tenant['section_value']

            date_last_paid = tenant['date_last_paid']
            amount = request.form.get('amount')
            if amount:
                amount = int(amount)
                fields_to_update['amount'] = amount
                fields_to_update['available_amount'] = amount + tenant['available_amount'] - tenant['amount']
                balance = section_value - (amount + tenant['available_amount'] - tenant['amount'])
                # Create a payment receipt PDF file
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter)

                # QR Code Generation
                url = f'https://michmanagement.onrender.com//get_receipt?tenantId={tenantId}'
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=3,
                    border=4,
                )
                qr.add_data(url)
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                img.save(f'payment_receipt_qr_{login_data}.png')

                # Create the receipt details
                data = [
                    ['Rent Payment Receipt - ' + company['company_name'], ''],
                    ['Receipt for:', tenant['tenantName']],
                    ['Property Name:', tenant['propertyName']],
                    ['Payment Type:', tenant['payment_type']],
                    ['Amount Paid:', f"{tenant['currency']} {amount}"],
                    ['Payment Mode:', tenant['payment_mode']],
                    ['Month Paid for:', tenant['months_paid']],
                    ['Date Paid:', date_last_paid.strftime('%Y-%m-%d')],
                    ['Balance:', f"{tenant['currency']} {balance}"],
                    ['Prepared by:', company['name']]
                ]

                # Create a table with the receipt details
                table = Table(data)

                # Add a table style
                table.setStyle(TableStyle([
                    ('SPAN', (0, 0), (1, 0)),  # Merge the first row
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),

                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),

                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0,0), (-1,-1), 1, colors.black),
                    ('FONTNAME', (1, -1), (1, -1), 'Helvetica-Oblique')  # Make the last cell on the last row italic
                ]))

                # Load your QR code image
                qr_code_img = f'payment_receipt_qr_{login_data}.png'
                qr_code = Image(qr_code_img)
                qr_code.hAlign = 'CENTER'

                # Add the QR code image to the elements list before building the PDF
                elements = [table, qr_code]
                doc.build(elements)

                # Get the PDF data and encode it as base64
                pdf_data = buffer.getvalue()
                buffer.close()
                payment_receipt_base64 = base64.b64encode(pdf_data).decode()

                # Delete the QR code image file
                os.remove(f'payment_receipt_qr_{login_data}.png')

                fields_to_update['payment_receipt'] = payment_receipt_base64

                # Create the email message
                if send_emails is not None:
                    msg = Message('Rent Payment Receipt-Mich Manage', 
                                sender='michmanage@outlook.com', 
                                recipients=[tenant['tenantEmail']])
                    msg.html = f"""
                    <html>
                    <body>
                    <p>Dear {tenant['tenantName']},</p>
                    <p>Please find attached your payment receipt for {tenant['months_paid']} {date_last_paid.year}.</p>
                    <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                    <p>Best Regards,</p>
                    <p>Mich Manage</p>
                    </body>
                    </html>
                    """

                    # Attach the PDF receipt to the email
                    msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                    # Send the email
                    thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                    thread.start()

            payment_mode = request.form.get('payment_mode')
            if payment_mode:
                fields_to_update['payment_mode'] = payment_mode

            months_paid = request.form.get('months_paid')
            if months_paid:
                fields_to_update['months_paid'] = months_paid

            date_last_paid = request.form.get('date_last_paid')
            if date_last_paid:
                date_last_paid = datetime.strptime(date_last_paid, '%Y-%m-%d')
                fields_to_update['date_last_paid'] = date_last_paid

            if amount and section_value:
                if amount < section_value:
                    payment_completion = 'Partial'
                elif amount > section_value:
                    flash('Amount entered should not exceed section value', 'error')
                    return redirect(url_for('propertyManagement_route.edit', tenantId=tenantId))
                else:
                    payment_completion = 'Full'
                fields_to_update['payment_completion'] = payment_completion

            db.tenants.update_one({'_id': ObjectId(tenantId)},{'$set': fields_to_update})

            flash('Tenant was successfully edited', 'success')
            tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenant['tenantEmail'], 'propertyName': tenant['propertyName']})
            if tenant_user:
                db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                db.userNotifications.insert_one({
                    'category': 'payment',
                    'user': tenant_user["_id"],
                    'notification': f"New payment recorded by manager {login_data}",
                    'timestamp': datetime.utcnow()
                })
            db.audit_logs.insert_one({'user': login_data, 'Activity': 'Edit tenant data', 'tenantEmail':tenant['tenantEmail'], 'timestamp': datetime.now()})
            return redirect('/update-tenant-info')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
        
###########VIEW TENANT RECEIPT###############
@propertyManagement.route('/view-receipt/<tenant_email>/<property_name>/<selected_section>', methods=["GET"])
def view_receipt(tenant_email, property_name, selected_section):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    # Retrieve the tenant document using tenant_id
    tenant = db.tenants.find_one({'username': login_data, 'propertyName': property_name, 'selected_section': selected_section, 'tenantEmail': tenant_email})

    if tenant is not None and 'payment_receipt' in tenant:
        # Convert the base64 string back to bytes
        payment_receipt = base64.b64decode(tenant['payment_receipt'])

        # Create a BytesIO object from the PDF data
        pdf_io = io.BytesIO(payment_receipt)

        # Create the file name
        file_name = f"{property_name}_{selected_section}_{{tenant['months_paid']}}_{{tenant['year']}}.pdf"

        # Create a custom response
        response = make_response(pdf_io.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={file_name}'

        return response

    else:
        return "No receipt found for this tenant", 404
                
#############ADD TENANT####################
@propertyManagement.route('/add-tenant', methods=["POST"])
def add_tenant():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            send_emails = db.send_emails.find_one({'emails': "yes"},{'emails': 1})

            tenantName = request.form.get('tenantName')
            gender = request.form.get('gender')
            household_size = request.form.get('household_size')
            tenantEmail = request.form.get('tenantEmail')
            tenantPhone = request.form.get('tenantPhone')
            propertyName = request.form.get('propertyName')
            selected_section = request.form.get('selected_section')
            section_value = request.form.get('section_value')
            section_value = int(section_value)
            payment_type = request.form.get('payment_type')
            amount = request.form.get('amount')
            payment_completion = request.form.get('payment_completion')
            amount = int(amount)
            currency = request.form.get('currency')
            payment_mode = request.form.get('payment_mode')
            months_paid = request.form.get('months_paid')
            date_last_paid = request.form.get('date_last_paid')
            date_last_paid = datetime.strptime(date_last_paid, '%Y-%m-%d')

            balance = section_value - amount

            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})

            # Calculate the number of full months the payment covers
            num_full_months = amount // section_value
            receipt_month = months_paid
            if num_full_months > 1:
                number_of_months = num_full_months
                receipt_month = f"{number_of_months} months starting from {months_paid}"
                balance = 0
            # Create a payment receipt PDF file
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)

            # QR Code Generation
            url = f'https://michmanagement.onrender.com//get_receipt?tenantEmail={tenantEmail}&propertyName={propertyName}&selected_section={selected_section}&months_paid={months_paid}&year={date_last_paid.year}'
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=3,
                border=4,
            )
            qr.add_data(url)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            img.save(f'payment_receipt_qr_{login_data}.png')

            # Create the receipt details
            data = [
                ['Rent Payment Receipt - ' + company['company_name'], ''],
                ['Receipt for:', tenantName],
                ['Property Name:', propertyName],
                ['Payment Type:', payment_type],
                ['Amount Paid:', f"{currency} {amount}"],
                ['Payment Mode:', payment_mode],
                ['Month Paid for:', receipt_month],
                ['Date Paid:', date_last_paid.strftime('%Y-%m-%d')],
                ['Balance:', f"{currency} {balance}"],
                ['Prepared by:', company['name']]
            ]

            # Create a table with the receipt details
            table = Table(data)

            # Add a table style
            table.setStyle(TableStyle([
                ('SPAN', (0, 0), (1, 0)),  # Merge the first row
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),

                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),

                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('FONTNAME', (1, -1), (1, -1), 'Helvetica-Oblique')  # Make the last cell on the last row italic
            ]))

            # Load your QR code image
            qr_code_img = f'payment_receipt_qr_{login_data}.png'
            qr_code = Image(qr_code_img)
            qr_code.hAlign = 'CENTER'

            # Add the QR code image to the elements list before building the PDF
            elements = [table, qr_code]
            doc.build(elements)

            # Get the PDF data and encode it as base64
            pdf_data = buffer.getvalue()
            buffer.close()
            payment_receipt_base64 = base64.b64encode(pdf_data).decode()

            # Delete the QR code image file
            os.remove(f'payment_receipt_qr_{login_data}.png')

            is_manager = db.managers.find_one({'manager_email': company['email'], 'name':company['company_name']})
            num_tenants = db.tenants.count_documents({'company_name': company['company_name']})
            if is_manager['amount_per_month'] == 70000 and num_tenants>=50:
                flash('Maximum number of tenants is reached', 'error')
                return redirect('/load-dashboard-page')
            elif is_manager['amount_per_month'] == 100000 and num_tenants>=100:
                flash('Maximum number of tenants is reached', 'error')
                return redirect('/load-dashboard-page')
            elif is_manager['amount_per_month'] == 150000 and num_tenants>=200:
                flash('Maximum number of tenants is reached', 'error')
                return redirect('/load-dashboard-page')
            else:
                section_exists = db.tenants.find_one({'company_name': company['company_name'], 'propertyName': propertyName, 'selected_section': selected_section})
                if amount < section_value:
                    payment_completion = 'Partial'
                elif amount > section_value:
                    # Get the starting month number from the form data
                    starting_month = list(calendar.month_name).index(months_paid)

                    # Calculate the remaining amount after the full months
                    remaining_amount = amount % section_value

                    # Create a list to store the data for each month
                    monthly_data = []

                    # Add the data for the full months
                    for i in range(num_full_months):
                        month_number = (starting_month + i - 1) % 12 + 1
                        year = date_last_paid.year + (starting_month + i - 1) // 12
                        monthly_data.append({
                            'username': login_data,
                            'company_name': company['company_name'],
                            'tenantName': tenantName,
                            'gender': gender,
                            'household_size': household_size,
                            'tenantEmail': tenantEmail,
                            'tenantPhone': tenantPhone,
                            'propertyName': propertyName,
                            'selected_section': selected_section,
                            'section_value': section_value,
                            'payment_type': payment_type,
                            'amount': section_value,
                            'payment_mode': payment_mode,
                            'months_paid': calendar.month_name[month_number],
                            'year': year,
                            'available_amount': section_value,
                            'payment_completion': 'Full',
                            'currency': currency,
                            'date_last_paid': date_last_paid,
                            'payment_status': '',
                            'status': '',
                            'payment_receipt': payment_receipt_base64
                        })

                    # If there is a remaining amount, add the data for the partial month
                    if remaining_amount > 0:
                        month_number = (starting_month + num_full_months - 1) % 12 + 1
                        year = date_last_paid.year + (starting_month + num_full_months - 1) // 12
                        monthly_data.append({
                            'username': login_data,
                            'company_name': company['company_name'],
                            'tenantName': tenantName,
                            'gender': gender,
                            'household_size': household_size,
                            'tenantEmail': tenantEmail,
                            'tenantPhone': tenantPhone,
                            'propertyName': propertyName,
                            'selected_section': selected_section,
                            'section_value': section_value,
                            'payment_type': payment_type,
                            'amount': remaining_amount,
                            'payment_mode': payment_mode,
                            'months_paid': calendar.month_name[month_number],
                            'year': year,
                            'available_amount': remaining_amount,
                            'payment_completion': 'Partial',
                            'currency': currency,
                            'date_last_paid': date_last_paid,
                            'payment_status': '',
                            'status': '',
                            'payment_receipt': payment_receipt_base64
                        })

                    # Now you can store the data for each month separately
                    for i, data in enumerate(monthly_data):
                        if i < len(monthly_data) - 1:  # If it's not the last record
                            # Store in 'old_tenant_data'
                            db.old_tenant_data.insert_one(data)
                        else:  # If it's the last record
                            # Store in 'tenants', regardless of whether it's fully paid or not
                            db.tenants.insert_one(data)
                    
                    tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                    if tenant_user:
                        db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                        db.userNotifications.insert_one({
                            'category': 'payment',
                            'user': tenant_user["_id"],
                            'notification': f"New payment recorded by manager {login_data}",
                            'timestamp': datetime.utcnow()
                        })
                    # Create the email message
                    if send_emails is not None:
                        msg = Message('Rent Payment Receipt-Mich Manage', 
                                    sender='michmanage@outlook.com', 
                                    recipients=[tenantEmail])
                        msg.html = f"""
                        <html>
                        <body>
                        <p>Dear {tenantName},</p>
                        <p>Please find attached your payment receipt for {receipt_month} {date_last_paid.year}.</p>
                        <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                        <p>Best Regards,</p>
                        <p>Mich Manage</p>
                        </body>
                        </html>
                        """

                        # Attach the PDF receipt to the email
                        msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                        # Send the email
                        thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                        thread.start()
                    db.audit_logs.insert_one({'user': login_data, 'Activity': 'Add tenant data', 'tenantName': tenantName, 'timestamp': datetime.now()})
                    flash('Tenant was successfully added', 'success')
                    return redirect('/load-dashboard-page')
                else:
                    payment_completion = 'Full'
                if section_exists is None:
                    tenant_details = {'username': login_data,
                                        'company_name': company['company_name'],
                                        'tenantName': tenantName,
                                        'gender': gender,
                                        'household_size': household_size,
                                        'tenantEmail': tenantEmail,
                                        'tenantPhone': tenantPhone,
                                        'propertyName': propertyName,
                                        'selected_section': selected_section,
                                        'section_value': section_value,
                                        'payment_type': payment_type,
                                        'amount': amount,
                                        'payment_mode': payment_mode,
                                        'months_paid': months_paid,
                                        'year': date_last_paid.year,
                                        'available_amount': amount,
                                        'payment_completion': payment_completion,
                                        'currency': currency,
                                        'date_last_paid': date_last_paid,
                                        'payment_status': '',
                                        'status': '',
                                        'payment_receipt': payment_receipt_base64}
                    
                    db.tenants.insert_one(tenant_details)
                    tenant_user = db.tenant_user_accounts.find_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
                    if tenant_user:
                        db.userNotifications.create_index([("timestamp", ASCENDING)], expireAfterSeconds=20)
                        db.userNotifications.insert_one({
                            'category': 'payment',
                            'user': tenant_user["_id"],
                            'notification': f"New payment recorded by manager {login_data}",
                            'timestamp': datetime.utcnow()
                        })
                    # Create the email message
                    if send_emails is not None:
                        msg = Message('Rent Payment Receipt-Mich Manage', 
                                    sender='michmanage@outlook.com', 
                                    recipients=[tenantEmail])
                        msg.html = f"""
                        <html>
                        <body>
                        <p>Dear {tenantName},</p>
                        <p>Please find attached your payment receipt for {months_paid} {date_last_paid.year}.</p>
                        <p><b><a href="https://michmanagement.onrender.com//tenant%20login%20page">Login</a></b></p>
                        <p>Best Regards,</p>
                        <p>Mich Manage</p>
                        </body>
                        </html>
                        """

                        # Attach the PDF receipt to the email
                        msg.attach("Rent Payment Receipt.pdf", "application/pdf", pdf_data)

                        # Send the email
                        thread = threading.Thread(target=send_async_email, args=[current_app._get_current_object(), msg])
                        thread.start()
                    db.audit_logs.insert_one({'user': login_data, 'Activity': 'Add tenant data', 'tenantName':tenantName, 'timestamp': datetime.now()})
                    flash('Tenant was successfully added', 'success')
                    return redirect('/load-dashboard-page')
                else:
                    flash('Section is already assigned', 'error')
                    return redirect('/load-dashboard-page')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

########DELETE TENANT################
@propertyManagement.route('/delete_tenant/<tenantEmail>/<propertyName>/<selected_section>')
def delete_tenant(tenantEmail, propertyName, selected_section):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
            tenants = db.tenants.find_one({'company_name': company['company_name'], 'tenantEmail': tenantEmail, 'propertyName': propertyName, 'selected_section': selected_section})
            # Remove the _id field
            if '_id' in tenants:
                del tenants['_id']
            db.old_tenant_data.insert_one(tenants)
            db.old_tenant_data.update_one({'company_name': company['company_name'], 'tenantEmail': tenantEmail, 'propertyName': propertyName, 'selected_section': selected_section}, {'$set': {'status': 'deleted'}})
            db.tenants.delete_one({'company_name': company['company_name'], 'tenantEmail': tenantEmail, 'propertyName': propertyName, 'selected_section': selected_section})
            db.tenant_user_accounts.delete_one({'tenantEmail': tenantEmail, 'propertyName': propertyName})
            db.audit_logs.insert_one({'user': login_data, 'Activity': 'Delete tenant', 'tenantName': tenants['tenantName'], 'timestamp': datetime.now()})
            return redirect('/update-tenant-info')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

#############MANAGER DOWNLOAD DATA######################
@propertyManagement.route('/download', methods=["POST"])
def download():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            startdate_on_str = request.form.get("startdate")
            enddate_on_str = request.form.get("enddate")
            startdate = datetime.strptime(startdate_on_str, '%Y-%m-%d')
            enddate = datetime.strptime(enddate_on_str, '%Y-%m-%d')

            # Fetch company information
            company = db.registered_managers.find_one({'username': login_data}, {'_id': 0, 'createdAt': 0, 'code': 0, 'address': 0, 'password': 0, 'auth': 0, 'dark_mode': 0})
            if not company:
                flash('Company not found', 'error')
                return redirect('/manager login page')
            
            is_manager = db.managers.find_one({'manager_email': company['email'], 'name':company['company_name']})
            if is_manager is None:
                user_query = {'username': login_data, 'company_name': company['company_name'], 'date_last_paid': {'$gte': startdate, '$lte': enddate}}
            else:
                user_query = {'company_name': company['company_name'], 'date_last_paid': {'$gte': startdate, '$lte': enddate}}
            
            projection = {'payment_receipt': 0, '_id': 0, 'marital_status': 0, 'age': 0, 'available_amount': 0, 'payment_completion': 0,
                        'currency': 0, 'payment_status': 0, 'status': 0, 'household_size': 0}

            # Fetch tenant data
            current_tenant_data = list(db.tenants.find(user_query, projection))
            old_tenant_data = list(db.old_tenant_data.find(user_query, projection))

            if not (current_tenant_data or old_tenant_data):
                flash('No tenant data found', 'error')
                return redirect('/load-dashboard-page')
            
            # Combine data
            combined_data = current_tenant_data + old_tenant_data

            # Define column names and month order
            new_column_names = {
                'username': 'Property manager',
                'company_name': 'Company',
                'tenantName': 'Tenant name',
                'gender': 'Gender',
                'tenantEmail': 'Tenant email',
                'tenantPhone': 'Tenant phone',
                'propertyName': 'Property name',
                'selected_section': 'Section',
                'section_value': 'Section value',
                'payment_type': 'Payment type',
                'amount': 'Amount paid',
                'payment_mode': 'Payment mode',
                'months_paid': 'Month paid',
                'year': 'Year',
                'date_last_paid': 'Date paid',
            }
            
            month_order = {'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6, 'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12}
            
            # Create an Excel workbook
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Tenants"

            # Write column headers
            headers = list(new_column_names.values())
            for col_idx, col_name in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Write data rows
            for r_idx, item in enumerate(combined_data, start=2):
                for c_idx, key in enumerate(new_column_names.keys(), start=1):
                    value = item.get(key, '')
                    if key == 'months_paid' and value in month_order:
                        value = month_order[value]
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Protect the Excel file with a password
            file_password = generate_file_password()
            ws.protection.set_password(file_password)

            existing_password = db.file_passwords.find_one({'username': login_data, 'detail': 'Tenant data file'})
            if existing_password:
                db.file_passwords.delete_one({'username': login_data, 'detail': 'Tenant data file'})
            db.file_passwords.insert_one({'username': login_data, 'password': file_password, 'detail': 'Tenant data file'})

            # Save the workbook to a BytesIO buffer
            output.seek(0)
            wb.save(output)
            wb.close()

            # Set the buffer position to the beginning
            output.seek(0)
            protected_data = output.read()

            # Create a response with the protected Excel file
            response = make_response(protected_data)
            response.headers['Content-Disposition'] = f"attachment; filename={company['company_name']}_{startdate_on_str}_{enddate_on_str}.xlsx"
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            # Optionally, save the password to the database (if needed)
            existing_password = db.file_passwords.find_one({'username': login_data, 'detail': 'Tenant data file'})
            if existing_password:
                db.file_passwords.delete_one({'username': login_data, 'detail': 'Tenant data file'})
            db.file_passwords.insert_one({'username': login_data, 'password': file_password, 'detail': 'Tenant data file'})

            return response
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
        
#####FILE PASSWORDS
@propertyManagement.route('/view-file-passwords')
def view_file_passwords():
    db, fs = get_db_and_fs()
    username = session.get('login_username')
    if username is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            company = db.registered_managers.find_one({'username': username})
            if 'dp' in company:
                dp_str = company['dp']
            else:
                dp_str = None
            
            file_passwords = list(db.file_passwords.find({'username': username}))
            if len(file_passwords)==0:
                flash('No encryption keys found', 'error')
                return redirect('/load-dashboard-page')
            else:
                found_passwords = []
                for password in file_passwords:
                    found_passwords.append(password)
            return render_template('file passwords.html', found_passwords=found_passwords, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

####MANAGE CONTRACTS
@propertyManagement.route('/manage-contracts')
def manage_contracts():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            if session.get('manage_contracts') == "no":
                flash("You do not have rights to manage contracts","error")
                return redirect("/load-dashboard-page")
            else:
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                is_manager = db.managers.find_one({'manager_email': company['email'], 'name':company['company_name']})
                ####CHECK IS LOGEDIN MANAGER HAS FULL RIGHTS
                if is_manager is None:
                    user_querry = {'username': login_data, 'company_name': company['company_name']}
                else:
                    user_querry = {'company_name': company['company_name']}

                contracts = list(db.contracts.find(user_querry))
                if len(contracts)==0:
                    flash('You are not managing any contracts!', 'error')
                    return redirect('/load-dashboard-page')
                else:
                    tenant_contracts = []
                    for contract in contracts:
                        end_date = contract['end_date']
                        now = datetime.now()
                        # Calculate the remaining period from now
                        remaining_seconds = int((end_date - now).total_seconds())
                        remaining_minutes, remaining_seconds = divmod(remaining_seconds, 60)
                        remaining_hours, remaining_minutes = divmod(remaining_minutes, 60)
                        remaining_days, remaining_hours = divmod(remaining_hours, 24)
                        remaining_days += 1

                        if remaining_days < 0:
                            contract['remaining'] = 'Expired'
                        elif remaining_days < 7:
                            contract['remaining'] = f"In {remaining_days} days, {remaining_hours} hours, and {remaining_minutes} minutes"
                        elif 7 <= remaining_days < 30:
                            weeks, days = divmod(remaining_days, 7)
                            contract['remaining'] = f"In {weeks} weeks, {days} days, {remaining_hours} hours, and {remaining_minutes} minutes"
                        elif 30 <= remaining_days < 365:
                            months, days = divmod(remaining_days, 30)
                            contract['remaining'] = f"In {months} months, {days} days, {remaining_hours} hours, and {remaining_minutes} minutes"
                        else:
                            years, days = divmod(remaining_days, 365)
                            contract['remaining'] = f"In {years} years, {days} days, {remaining_hours} hours, and {remaining_minutes} minutes"
                        tenant_contracts.append(contract)
                    return render_template('manage contracts.html', tenant_contracts=tenant_contracts, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
        
@propertyManagement.route('/upload-contract-page')
def upload_contract_page():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            if session.get('manage_contracts') == "no":
                flash("You do not have rights to add contracts","error")
                return redirect("/load-dashboard-page")
            else:
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                is_manager = db.managers.find_one({'manager_email': company['email'], 'name':company['company_name']})
                ####CHECK IS LOGEDIN MANAGER HAS FULL RIGHTS
                if is_manager is None:
                    user_querry = {'username': login_data, 'company_name': company['company_name']}
                else:
                    user_querry = {'company_name': company['company_name']}
                
                tenants = list(db.tenants.find(user_querry))

                if len(tenants) == 0:
                    flash('No tenant data found', 'error')
                    return redirect('/load-dashboard-page')
                else:
                    tenant_names = []
                    for tenant in tenants:
                        tenant_names.append(tenant['tenantName'])
                return render_template('add contracts.html',tenant_names=tenant_names, dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@propertyManagement.route('/upload-contract', methods=['POST'])
def upload_contract():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            if session.get('manage_contracts') == "no":
                flash("You do not have rights to add contracts","error")
                return redirect("/load-dashboard-page")
            else:
                if 'contract_document' not in request.files:
                    flash("No file part", 'error')
                    return redirect('/upload-contract-page')
                file = request.files['contract_document']
                if file.filename == '':
                    flash("No file selected", 'error')
                    return redirect('/upload-contract-page')
                if file:
                    filename = secure_filename(file.filename)
                    content_type = file.content_type
                    file_id = fs.put(file.read(), filename=filename, content_type=content_type)
                    company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                    receiver = request.form.get('receiver')
                    start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d')
                    end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d')

                    contract = {
                        'username': login_data,
                        'company_name': company['company_name'],
                        'file_id': file_id,
                        'issuer': company['name'],
                        'receiver': receiver,
                        'start_date': start_date,
                        'end_date': end_date
                    }

                    db.contracts.insert_one(contract)
                    db.audit_logs.insert_one({'user': login_data, 'Activity': 'Add new contract', 'file_id': file_id, 'timestamp': datetime.now()})
                    flash("Contract was uploaded successfully", 'success')
                    return redirect('/upload-contract-page')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

########DELETE CONTRACTS################
@propertyManagement.route('/delete-contract/<contractID>')
def delete_contract(contractID):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            if session.get('manage_contracts') == "no":
                flash("You do not have rights to delete contracts","error")
                return redirect("/load-dashboard-page")
            else:
                contract = db.contracts.find_one({'_id': ObjectId(contractID)})
                # Remove the _id field
                if '_id' in contract:
                    del contract['_id']
                db.old_contracts.insert_one(contract)
                db.contracts.delete_one({'_id': ObjectId(contractID)})
                db.audit_logs.insert_one({'user': login_data, 'Activity': 'Delete contract', 'contractID':contractID, 'timestamp': datetime.now()})
                return redirect('/manage-contracts')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

##UPDATE CONTRACTS
@propertyManagement.route('/update-contract/<contractID>/<company_name>/<receiver>')
def selected_contract(contractID, company_name, receiver):
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            if session.get('manage_contracts') == "no":
                flash("You do not have rights to update contracts","error")
                return redirect("/load-dashboard-page")
            else:
                company = db.registered_managers.find_one({'username': login_data},{'_id':0,'createdAt':0,'code':0,'address':0,'password':0,'auth':0,'dark_mode':0})
                if 'dp' in company:
                    dp_str = company['dp']
                else:
                    dp_str = None
                return render_template('update contract.html',contractID=contractID,company_name=company_name,receiver=receiver,dp=dp_str)
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')

@propertyManagement.route('/updated-contract', methods=['POST'])
def updated_contract():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')
    if login_data is None:
        flash('Login first', 'error')
        return redirect('/manager login page')
    else:
        account_type = session.get('account_type')
        if account_type == 'Property Management':
            if session.get('manage_contracts') == "no":
                flash("You do not have rights to update contracts","error")
                return redirect("/load-dashboard-page")
            else:
                if 'contract_document' not in request.files:
                    flash("No file part", 'error')
                    return redirect('/manage-contracts')
                file = request.files['contract_document']
                if file.filename == '':
                    flash("No file selected", 'error')
                    return redirect('/manage-contracts')
                if file:
                    filename = secure_filename(file.filename)
                    content_type = file.content_type
                    file_id = fs.put(file.read(), filename=filename, content_type=content_type)
                    end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d')
                    contractID = request.form.get('contractID')

                    contract = db.contracts.find_one({'_id': ObjectId(contractID)})
                    if contract:
                        # If a contract exists, update the document and end date
                        fs.delete(contract['file_id'])
                        db.contracts.update_one(
                            {'_id': ObjectId(contractID)},
                            {'$set': {'file_id': file_id, 'end_date': end_date}}
                        )
                        db.audit_logs.insert_one({'user': login_data, 'Activity': 'Update contract', 'file_id': file_id, 'timestamp': datetime.now()})
                        flash("Contract was updated successfully", 'success')
                        return redirect('/manage-contracts')
                    else:
                        flash("Contract was not found", 'error')
                        return redirect('/manage-contracts')
        else:
            flash('Your session expired or does not exist', 'error')
            return redirect('/manager login page')
            
@propertyManagement.route('/download-contract/<fileID>')
def download_contract(fileID):
    db, fs = get_db_and_fs()
    file = fs.get(ObjectId(fileID))
    
    # Create a BytesIO buffer to zip the file
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        zip_file.writestr(file.filename, file.read())
    
    zip_buffer.seek(0)
    zip_data = zip_buffer.getvalue()
    
    response = make_response(zip_data)
    response.headers['Content-Disposition'] = f'attachment; filename={file.filename}.zip'
    response.headers['Content-Type'] = 'application/zip'
    
    return response

@propertyManagement.route('/notifications')
def notifications():
    db, fs = get_db_and_fs()
    login_data = session.get('login_username')

    manager_account = db.registered_managers.find_one({'username': login_data}, {'_id':0,'createdAt':0,'code':0,'phone_number':0,'address':0,'registered_on':0,'password':0,'auth':0,'dp':0,'dark_mode':0,'password':0})
    if manager_account:
        fields = ['add_properties', 'add_tenants', 'update_tenant', 'edit_tenant', 'manage_contracts', 'add_stock', 'update_stock',
                      'update_sales','inhouse','view_stock_info','view_revenue','view_sales','system_selling_price','point_of_sale',
                      'view_finance_dashboard','add_new_finance_account','update_finance_account','view_finance','edit_finance',
                      'delete_finance']
        
        is_manager = db.managers.find_one({'manager_email': manager_account['email'], 'name':manager_account['company_name']})
        if is_manager:
            for field in fields:
                value = manager_account.get(field)
                if value is not None:
                    session[field] = value
        else:
            for field in fields:
                if field in manager_account:
                    value = manager_account.get(field)
                    if value is not None:
                        session[field] = value
                else:
                    session[field] = "no"

        account_type = manager_account.get('account_type')
        if account_type is None:
            manager_type = db.managers.find_one({'name': manager_account['company_name']})
            account_type = manager_type['account_type']
            # Remove any empty strings from the list
            account_type = [atype for atype in account_type if atype]
            if 'Enterprise Resource Planning' in account_type and len(account_type) == 1:
                session['account_type'] = 'Enterprise Resource Planning'
            elif 'Property Management' in account_type and len(account_type) == 1:
                session['account_type'] = 'Property Management'
            elif 'Accounting' in account_type and len(account_type) == 1:
                session['account_type'] = 'Accounting'
        else:
            if account_type == 'Property Management':
                session['account_type'] = 'Property Management'
            elif account_type == 'Enterprise Resource Planning':
                session['account_type'] = 'Enterprise Resource Planning'
            elif 'Accounting' in account_type and len(account_type) == 1:
                session['account_type'] = 'Accounting'

    # Get the last seen timestamp from the session
    last_seen_timestamp = session.get('last_seen_timestamp', datetime.min)

    # Get the list of viewed notification IDs from the session
    viewed_notifications = session.get('viewed_notifications', [])

    # Fetch notifications with a timestamp greater than the last seen timestamp
    notifications_cursor = db.userNotifications.find({
        'user': login_data,
        'timestamp': {'$gt': last_seen_timestamp},
        '_id': {'$nin': [ObjectId(id) for id in viewed_notifications]}  # Exclude already viewed notifications
    }, {'_id': 1, 'notification': 1, 'timestamp': 1})

    # Convert cursor to list
    notifications_list = list(notifications_cursor)

    # Prepare the response with only new notifications
    notifications_to_send = [
        {
            'notification': notification['notification'],
            'timestamp': notification['timestamp'].isoformat()
        }
        for notification in notifications_list
    ]
    
    if notifications_to_send:
        # Update the last seen timestamp to the maximum timestamp of the fetched notifications
        new_last_seen_timestamp = max(notification['timestamp'] for notification in notifications_list)
        session['last_seen_timestamp'] = new_last_seen_timestamp

        # Update the list of viewed notifications
        new_viewed_notifications = [str(notification['_id']) for notification in notifications_list]
        session['viewed_notifications'] = viewed_notifications + new_viewed_notifications
    
    # Prepare the response with only new notifications
    notifications_list = [notification['notification'] for notification in notifications_to_send]

    return jsonify(notifications_list)